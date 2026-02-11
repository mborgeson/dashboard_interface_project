"""
Structural fingerprinting of Excel UW model files.

Produces a SheetFingerprint for each sheet (name, dimensions, header labels,
column-A labels, populated cell count) and a FileFingerprint per file
(metadata + per-sheet fingerprints + population classification).

Fingerprinting runs in a ProcessPoolExecutor for parallelism since it
involves CPU-bound workbook parsing.
"""

import hashlib
import io
from concurrent.futures import as_completed
from dataclasses import asdict, dataclass, field
from pathlib import Path
from typing import Any

import structlog

logger = structlog.get_logger().bind(component="fingerprint")


@dataclass
class SheetFingerprint:
    """Structural fingerprint for a single Excel sheet."""

    name: str
    row_count: int = 0
    col_count: int = 0
    header_labels: list[str] = field(default_factory=list)
    col_a_labels: list[str] = field(default_factory=list)
    populated_cell_count: int = 0

    @property
    def signature(self) -> str:
        """Deterministic signature from sheet structure."""
        parts = [
            self.name,
            str(self.row_count),
            str(self.col_count),
            "|".join(sorted(self.header_labels)),
            "|".join(sorted(self.col_a_labels)),
        ]
        return hashlib.md5(";".join(parts).encode(), usedforsecurity=False).hexdigest()

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)

    @classmethod
    def from_dict(cls, d: dict[str, Any]) -> "SheetFingerprint":
        return cls(**{k: v for k, v in d.items() if k in cls.__dataclass_fields__})


@dataclass
class FileFingerprint:
    """Structural fingerprint for an entire Excel file."""

    file_path: str
    file_name: str
    file_size: int = 0
    content_hash: str = ""
    sheet_count: int = 0
    sheet_signatures: list[str] = field(default_factory=list)
    sheets: list[SheetFingerprint] = field(default_factory=list)
    total_populated_cells: int = 0
    population_status: str = "unknown"  # populated, sparse, empty

    def to_dict(self) -> dict[str, Any]:
        d = asdict(self)
        d["sheets"] = [
            s.to_dict() if isinstance(s, SheetFingerprint) else s for s in self.sheets
        ]
        return d

    @classmethod
    def from_dict(cls, d: dict[str, Any]) -> "FileFingerprint":
        sheets_data = d.pop("sheets", [])
        sheets = [
            SheetFingerprint.from_dict(s) if isinstance(s, dict) else s
            for s in sheets_data
        ]
        fp = cls(**{k: v for k, v in d.items() if k in cls.__dataclass_fields__})
        fp.sheets = sheets
        return fp

    @property
    def combined_signature(self) -> str:
        """Combined signature of all sheets (includes labels — exact match)."""
        return "|".join(sorted(self.sheet_signatures))

    @property
    def sheet_name_key(self) -> str:
        """Key based on sorted sheet names only (ignores deal-specific labels).

        Used for initial clustering in the grouping algorithm. Files from
        the same template share sheet names (e.g. "Summary", "Cash Flow")
        but differ in deal-specific data within those sheets.
        """
        return "|".join(sorted(s.name for s in self.sheets))


def fingerprint_file(
    file_path: str,
    file_content: bytes | None = None,
    empty_threshold: int = 20,
) -> FileFingerprint:
    """
    Compute structural fingerprint for a single Excel file.

    Args:
        file_path: Path to Excel file.
        file_content: Optional bytes content (avoids disk read).
        empty_threshold: Minimum populated cells to be classified as "populated".

    Returns:
        FileFingerprint with per-sheet structural data.
    """
    path = Path(file_path)
    file_ext = path.suffix.lower()

    # Read content if not provided
    if file_content is None:
        if not path.exists():
            return FileFingerprint(
                file_path=file_path,
                file_name=path.name,
                population_status="error",
            )
        file_content = path.read_bytes()

    content_hash = hashlib.sha256(file_content).hexdigest()
    file_size = len(file_content)

    sheets: list[SheetFingerprint] = []

    try:
        if file_ext == ".xlsb":
            sheets = _fingerprint_xlsb(file_path, file_content)
        else:
            sheets = _fingerprint_xlsx(file_path, file_content)
    except Exception as e:
        logger.warning("fingerprint_error", file=path.name, error=str(e))
        return FileFingerprint(
            file_path=file_path,
            file_name=path.name,
            file_size=file_size,
            content_hash=content_hash,
            population_status="error",
        )

    total_populated = sum(s.populated_cell_count for s in sheets)

    if total_populated >= empty_threshold:
        status = "populated"
    elif total_populated > 0:
        status = "sparse"
    else:
        status = "empty"

    return FileFingerprint(
        file_path=file_path,
        file_name=path.name,
        file_size=file_size,
        content_hash=content_hash,
        sheet_count=len(sheets),
        sheet_signatures=[s.signature for s in sheets],
        sheets=sheets,
        total_populated_cells=total_populated,
        population_status=status,
    )


def _fingerprint_xlsb(file_path: str, file_content: bytes) -> list[SheetFingerprint]:
    """Fingerprint sheets from an .xlsb file using pyxlsb."""
    import pyxlsb

    sheets: list[SheetFingerprint] = []
    wb = pyxlsb.open_workbook(io.BytesIO(file_content))

    try:
        for sheet_name in wb.sheets:
            try:
                with wb.get_sheet(sheet_name) as sheet:
                    header_labels: list[str] = []
                    col_a_labels: list[str] = []
                    populated = 0
                    max_row = 0
                    max_col = 0

                    for row_idx, row in enumerate(sheet.rows()):
                        if row:
                            max_row = row_idx + 1
                            for cell in row:
                                col_idx = cell.c if hasattr(cell, "c") else 0
                                max_col = max(max_col, col_idx + 1)
                                val = cell.v
                                if val is not None and str(val).strip():
                                    populated += 1
                                    val_str = str(val).strip()
                                    # First row = header labels
                                    if row_idx == 0:
                                        header_labels.append(val_str)
                                    # Column A (col_idx=0) = row labels
                                    if col_idx == 0:
                                        col_a_labels.append(val_str)

                        # Stop scanning after 200 rows for performance
                        if row_idx > 200:
                            break

                    sheets.append(
                        SheetFingerprint(
                            name=sheet_name,
                            row_count=max_row,
                            col_count=max_col,
                            header_labels=header_labels[:50],  # Limit for storage
                            col_a_labels=col_a_labels[:100],
                            populated_cell_count=populated,
                        )
                    )
            except Exception as e:
                logger.debug("sheet_fingerprint_error", sheet=sheet_name, error=str(e))
                sheets.append(SheetFingerprint(name=sheet_name))
    finally:
        wb.close()

    return sheets


def _fingerprint_xlsx(file_path: str, file_content: bytes) -> list[SheetFingerprint]:
    """Fingerprint sheets from an .xlsx/.xlsm file using openpyxl."""
    import warnings

    import openpyxl

    sheets: list[SheetFingerprint] = []

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        wb = openpyxl.load_workbook(
            io.BytesIO(file_content), read_only=True, data_only=True
        )

    try:
        for sheet_name in wb.sheetnames:
            try:
                ws = wb[sheet_name]
                header_labels: list[str] = []
                col_a_labels: list[str] = []
                populated = 0
                max_row = ws.max_row or 0
                max_col = ws.max_column or 0

                for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
                    for cell in row:
                        val = cell.value
                        if val is not None and str(val).strip():
                            populated += 1
                            val_str = str(val).strip()
                            if row_idx == 1:
                                header_labels.append(val_str)
                            if cell.column == 1:
                                col_a_labels.append(val_str)

                    # Stop scanning after 200 rows for performance
                    if row_idx > 200:
                        break

                sheets.append(
                    SheetFingerprint(
                        name=sheet_name,
                        row_count=max_row,
                        col_count=max_col,
                        header_labels=header_labels[:50],
                        col_a_labels=col_a_labels[:100],
                        populated_cell_count=populated,
                    )
                )
            except Exception as e:
                logger.debug("sheet_fingerprint_error", sheet=sheet_name, error=str(e))
                sheets.append(SheetFingerprint(name=sheet_name))
    finally:
        wb.close()

    return sheets


def fingerprint_files_parallel(
    file_paths: list[str],
    file_contents: dict[str, bytes] | None = None,
    max_workers: int = 4,
) -> list[FileFingerprint]:
    """
    Fingerprint multiple files in parallel.

    Uses ThreadPoolExecutor (not ProcessPoolExecutor) because pyxlsb
    workbook objects aren't picklable.

    Args:
        file_paths: List of file paths to fingerprint.
        file_contents: Optional dict mapping path -> bytes.
        max_workers: Maximum parallel workers.

    Returns:
        List of FileFingerprint objects.
    """
    from concurrent.futures import ThreadPoolExecutor

    results: list[FileFingerprint] = []

    if len(file_paths) <= 1:
        # No parallelism needed
        for fp in file_paths:
            content = file_contents.get(fp) if file_contents else None
            results.append(fingerprint_file(fp, content))
        return results

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        futures = {
            executor.submit(
                fingerprint_file,
                fp,
                file_contents.get(fp) if file_contents else None,
            ): fp
            for fp in file_paths
        }
        for future in as_completed(futures):
            try:
                results.append(future.result())
            except Exception as e:
                path = futures[future]
                logger.error("fingerprint_parallel_error", file=path, error=str(e))
                results.append(
                    FileFingerprint(
                        file_path=path,
                        file_name=Path(path).name,
                        population_status="error",
                    )
                )

    return results
