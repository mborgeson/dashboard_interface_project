"""
B&R Capital Dashboard - Excel Data Extractor

Extracts data from Excel underwriting models (.xlsb and .xlsx) using
cell mappings from the reference file.

Key features:
- Supports both .xlsb (binary) and .xlsx/.xlsm formats
- Proper 0-based indexing conversion for pyxlsb
- Comprehensive error handling with graceful NaN degradation
- Progress callbacks for API integration
"""

import io
import re
import numpy as np
import structlog
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any, Callable, Dict, List, Optional
from concurrent.futures import ThreadPoolExecutor, as_completed

import openpyxl
import pyxlsb

from .cell_mapping import CellMapping
from .error_handler import ErrorHandler


# Suppress openpyxl warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")


class ExcelDataExtractor:
    """
    Extracts data from Excel underwriting models (.xlsb and .xlsx).

    Uses cell mappings to extract specific values from sheets, with
    comprehensive error handling that returns np.nan for any failures.
    """

    def __init__(self, cell_mappings: Dict[str, CellMapping]):
        self.mappings = cell_mappings
        self.logger = structlog.get_logger().bind(component="ExcelDataExtractor")
        self.error_handler = ErrorHandler()

    def extract_from_file(
        self,
        file_path: str,
        file_content: Optional[bytes] = None,
        progress_callback: Optional[Callable[[int, int], None]] = None,
    ) -> Dict[str, Any]:
        """
        Extract all mapped values from an Excel file.

        Args:
            file_path: Path to Excel file (or filename if using file_content)
            file_content: Optional bytes content (e.g., from SharePoint download)
            progress_callback: Optional callback(current, total) for progress updates

        Returns:
            Dictionary with:
            - Extracted field values (field_name: value)
            - _file_path: Source file path
            - _extraction_timestamp: ISO timestamp
            - _extraction_errors: List of error details
            - _extraction_metadata: Processing statistics
        """
        start_time = datetime.now()

        # Reset error handler for this extraction
        self.error_handler.reset()

        extracted_data: Dict[str, Any] = {
            "_file_path": file_path,
            "_extraction_timestamp": datetime.now().isoformat(),
            "_extraction_errors": [],
        }

        # Validate file exists when using file path
        if file_content is None:
            path = Path(file_path)
            if not path.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

        # Determine file type and load workbook
        file_ext = Path(file_path).suffix.lower()

        if file_ext == ".xlsb":
            workbook = self._load_xlsb(file_path, file_content)
            is_xlsb = True
        else:
            workbook = self._load_xlsx(file_path, file_content)
            is_xlsb = False

        # Get available sheets for error reporting
        if is_xlsb:
            available_sheets = list(workbook.sheets)
        else:
            available_sheets = workbook.sheetnames

        self.logger.debug(
            "workbook_loaded",
            file_path=file_path,
            format="xlsb" if is_xlsb else "xlsx",
            sheets=available_sheets,
        )

        # Extract values for each mapping
        successful = 0
        failed = 0
        total = len(self.mappings)

        for i, (field_name, mapping) in enumerate(self.mappings.items()):
            try:
                value = self._extract_cell_value(
                    workbook,
                    mapping.sheet_name,
                    mapping.cell_address,
                    field_name,
                    is_xlsb,
                )
                extracted_data[field_name] = value

                if (
                    not np.isnan(value)
                    if isinstance(value, float)
                    else value is not None
                ):
                    successful += 1
                else:
                    failed += 1

            except Exception as e:
                extracted_data[field_name] = np.nan
                extracted_data["_extraction_errors"].append(
                    {
                        "field": field_name,
                        "sheet": mapping.sheet_name,
                        "cell": mapping.cell_address,
                        "error": str(e),
                    }
                )
                failed += 1

            # Progress callback
            if progress_callback and (i + 1) % 100 == 0:
                progress_callback(i + 1, total)

        # Close workbook if needed
        if not is_xlsb and hasattr(workbook, "close"):
            workbook.close()

        # Add extraction metadata
        duration = (datetime.now() - start_time).total_seconds()
        error_summary = self.error_handler.get_error_summary()

        extracted_data["_extraction_metadata"] = {
            "total_fields": total,
            "successful": successful,
            "failed": failed,
            "success_rate": round(successful / total * 100, 1) if total > 0 else 0,
            "duration_seconds": round(duration, 2),
            "error_summary": error_summary,
        }

        self.logger.info(
            "extraction_complete",
            file_path=file_path,
            successful=successful,
            failed=failed,
            duration=round(duration, 2),
        )

        return extracted_data

    def _load_xlsb(self, file_path: str, file_content: Optional[bytes] = None):
        """Load .xlsb file using pyxlsb"""
        if file_content:
            return pyxlsb.open_workbook(io.BytesIO(file_content))
        return pyxlsb.open_workbook(file_path)

    def _load_xlsx(self, file_path: str, file_content: Optional[bytes] = None):
        """Load .xlsx/.xlsm file using openpyxl"""
        if file_content:
            return openpyxl.load_workbook(
                io.BytesIO(file_content),
                data_only=True,  # Get calculated values, not formulas
                keep_vba=True,  # Preserve VBA for .xlsm files
            )
        return openpyxl.load_workbook(file_path, data_only=True, keep_vba=True)

    def _extract_cell_value(
        self,
        workbook,
        sheet_name: str,
        cell_address: str,
        field_name: str,
        is_xlsb: bool,
    ) -> Any:
        """
        Extract value from specific cell with comprehensive error handling.

        IMPORTANT: pyxlsb uses 0-based indexing while Excel uses 1-based.
        Excel D6 (row=6, col=4) → pyxlsb (row=5, col=3)
        """
        try:
            if is_xlsb:
                return self._extract_from_xlsb(
                    workbook, sheet_name, cell_address, field_name
                )
            else:
                return self._extract_from_xlsx(
                    workbook, sheet_name, cell_address, field_name
                )
        except Exception as e:
            return self.error_handler.handle_unknown_error(
                field_name, sheet_name, cell_address, str(e)
            )

    def _extract_from_xlsb(
        self, workbook, sheet_name: str, cell_address: str, field_name: str
    ) -> Any:
        """Extract value from .xlsb file with pyxlsb"""
        # Check sheet exists
        if sheet_name not in workbook.sheets:
            return self.error_handler.handle_missing_sheet(
                field_name, sheet_name, list(workbook.sheets)
            )

        # Parse cell address (e.g., "A1" -> row=0, col=0 in 0-based)
        clean_address = cell_address.replace("$", "").upper()
        match = re.match(r"^([A-Z]+)(\d+)$", clean_address)

        if not match:
            return self.error_handler.handle_invalid_cell_address(
                field_name,
                sheet_name,
                cell_address,
                "Invalid format - expected 'A1', 'B10', etc.",
            )

        col_str, row_str = match.groups()

        # Convert to 0-based indexing for pyxlsb
        target_row = int(row_str) - 1
        target_col = self._column_to_index(col_str)

        # Read from sheet
        with workbook.get_sheet(sheet_name) as sheet:
            for row in sheet.rows():
                for cell in row:
                    if cell.r == target_row and cell.c == target_col:
                        return self.error_handler.process_cell_value(
                            cell.v, field_name, sheet_name, cell_address
                        )

        # Cell not found (likely empty or outside bounds)
        return self.error_handler.handle_empty_value(
            field_name, sheet_name, cell_address
        )

    def _extract_from_xlsx(
        self, workbook, sheet_name: str, cell_address: str, field_name: str
    ) -> Any:
        """Extract value from .xlsx file with openpyxl"""
        # Check sheet exists
        if sheet_name not in workbook.sheetnames:
            return self.error_handler.handle_missing_sheet(
                field_name, sheet_name, workbook.sheetnames
            )

        sheet = workbook[sheet_name]

        # Clean cell address
        clean_address = cell_address.replace("$", "").upper()

        try:
            cell = sheet[clean_address]
            return self.error_handler.process_cell_value(
                cell.value, field_name, sheet_name, cell_address
            )
        except Exception:
            return self.error_handler.handle_cell_not_found(
                field_name, sheet_name, cell_address
            )

    def _column_to_index(self, col_str: str) -> int:
        """
        Convert Excel column letters to 0-based column index.

        A=0, B=1, ..., Z=25, AA=26, AB=27, ...
        """
        result = 0
        for char in col_str.upper():
            result = result * 26 + (ord(char) - ord("A") + 1)
        return result - 1  # Convert to 0-based


class BatchProcessor:
    """
    Processes multiple Excel files in batches with parallel execution.

    Provides progress tracking and error aggregation for API integration.
    """

    def __init__(
        self, extractor: ExcelDataExtractor, batch_size: int = 10, max_workers: int = 4
    ):
        self.extractor = extractor
        self.batch_size = batch_size
        self.max_workers = max_workers
        self.logger = structlog.get_logger().bind(component="BatchProcessor")

    def process_files(
        self,
        file_list: List[Dict[str, Any]],
        progress_callback: Optional[Callable[[int, int, str], None]] = None,
    ) -> Dict[str, Any]:
        """
        Process multiple files with parallel execution.

        Args:
            file_list: List of dicts with 'file_path', 'file_content' (optional),
                      'deal_name', 'deal_stage', 'modified_date'
            progress_callback: Optional callback(current, total, current_file)

        Returns:
            Dict with 'results', 'failed', 'summary'
        """
        total_files = len(file_list)
        processed_results: List[Dict[str, Any]] = []
        failed_files: List[Dict[str, Any]] = []

        self.logger.info(
            "starting_batch_processing",
            total_files=total_files,
            batch_size=self.batch_size,
            max_workers=self.max_workers,
        )

        start_time = datetime.now()

        # Process in batches
        for batch_start in range(0, total_files, self.batch_size):
            batch_end = min(batch_start + self.batch_size, total_files)
            batch = file_list[batch_start:batch_end]

            # Process batch with thread pool
            with ThreadPoolExecutor(max_workers=self.max_workers) as executor:
                future_to_file = {
                    executor.submit(self._process_single_file, file_info): file_info
                    for file_info in batch
                }

                for future in as_completed(future_to_file):
                    file_info = future_to_file[future]
                    file_path = file_info.get("file_path", "unknown")

                    try:
                        result = future.result()
                        processed_results.append(result)

                        if progress_callback:
                            progress_callback(
                                len(processed_results) + len(failed_files),
                                total_files,
                                file_path,
                            )

                    except Exception as e:
                        self.logger.error(
                            "file_processing_failed", file_path=file_path, error=str(e)
                        )
                        failed_files.append({"file_info": file_info, "error": str(e)})

        # Generate summary
        duration = (datetime.now() - start_time).total_seconds()

        summary = {
            "total_files": total_files,
            "processed": len(processed_results),
            "failed": len(failed_files),
            "success_rate": (
                round(len(processed_results) / total_files * 100, 1)
                if total_files > 0
                else 0
            ),
            "total_duration_seconds": round(duration, 2),
            "average_per_file": (
                round(duration / total_files, 2) if total_files > 0 else 0
            ),
            "failed_files": [f["file_info"].get("file_path") for f in failed_files],
        }

        self.logger.info("batch_processing_complete", **summary)

        return {
            "results": processed_results,
            "failed": failed_files,
            "summary": summary,
        }

    def _process_single_file(self, file_info: Dict[str, Any]) -> Dict[str, Any]:
        """Process a single file and add metadata"""
        file_path = file_info.get("file_path")
        file_content = file_info.get("file_content")

        # Extract data
        extracted_data = self.extractor.extract_from_file(file_path, file_content)

        # Add file metadata
        extracted_data.update(
            {
                "_deal_name": file_info.get("deal_name"),
                "_deal_stage": file_info.get("deal_stage"),
                "_file_modified_date": file_info.get("modified_date"),
            }
        )

        return extracted_data


# Custom Exception Classes
class ExtractionError(Exception):
    """Base exception for extraction errors"""

    pass


class FileAccessError(ExtractionError):
    """File access errors"""

    pass


class MappingError(ExtractionError):
    """Cell mapping errors"""

    pass
