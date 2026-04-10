"""
Template variant detection and cell address remapping.

Non-production UW model templates have key financial fields at different row
positions than the production template.  The group extraction pipeline
(group_pipeline.py) compensates via field_remaps.json, but the OneDrive /
SharePoint extraction path bypasses that pipeline and therefore extracts the
wrong cells for variant templates.

This module detects the template variant at extraction time by probing
specific cells in the "Assumptions (Summary)" sheet, then returns a
mapping of field_name -> corrected CellMapping so the caller can patch
its mappings dict before extraction.

Design decisions
----------------
* Detection is **cell-probe-based**, not filename-based, so it works for
  any file regardless of naming convention or origin.
* The set of known probe rows is loaded from ``field_remaps.json`` at
  import time (cached).  If the file is missing or malformed the
  detector gracefully returns no remaps.
* The module never modifies ``reference_mapper.py`` or
  ``cell_mapping.py`` (both on the Do-Not-Touch list).
"""

from __future__ import annotations

import copy
import io
import json
import re
import warnings
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl
import pyxlsb
from loguru import logger as _base_logger

logger = _base_logger.bind(component="variant_detector")

# Sheet where the variant-sensitive fields live
_ASSUMPTIONS_SHEET = "Assumptions (Summary)"

# Fields that require remapping in variant templates.
# Maps field_name -> (column_letter, prod_row, label_substring) where
# label_substring is checked in column A at the candidate row.
_VARIANT_FIELDS: dict[str, tuple[str, int, str]] = {
    "VACANCY_LOSS_YEAR_1_RATE": ("D", 216, "vacancy"),
    "PURCHASE_PRICE": ("D", 478, "purchase price"),
    "LOAN_AMOUNT": ("D", 359, "loan amount"),
}


@dataclass(frozen=True)
class VariantRemap:
    """A single field remap detected for a variant template."""

    field_name: str
    prod_cell: str
    variant_cell: str
    detected_row: int
    label_found: str


@dataclass
class VariantDetectionResult:
    """Result of variant detection for one file."""

    is_variant: bool
    remaps: list[VariantRemap]
    detection_method: str = "cell_probe"

    def to_dict(self) -> dict[str, Any]:
        return {
            "is_variant": self.is_variant,
            "detection_method": self.detection_method,
            "remap_count": len(self.remaps),
            "remaps": [
                {
                    "field": r.field_name,
                    "prod_cell": r.prod_cell,
                    "variant_cell": r.variant_cell,
                    "label_found": r.label_found,
                }
                for r in self.remaps
            ],
        }


def _load_known_candidate_rows(
    data_dir: Path | None = None,
) -> dict[str, list[int]]:
    """
    Build a map of field_name -> sorted list of known candidate rows
    from field_remaps.json.

    Falls back to a hardcoded set if the file is unavailable.
    """
    candidates: dict[str, set[int]] = {
        # Always include the production rows as candidates
        "VACANCY_LOSS_YEAR_1_RATE": {216, 45},
        "PURCHASE_PRICE": {478},
        "LOAN_AMOUNT": {359},
    }

    if data_dir is None:
        from app.core.config import settings

        data_dir = Path(settings.GROUP_EXTRACTION_DATA_DIR)

    remaps_path = data_dir / "field_remaps.json"
    if remaps_path.exists():
        try:
            remaps_data = json.loads(remaps_path.read_text())
            for _group_name, group_data in remaps_data.items():
                for field_name, remap_info in group_data.get("remaps", {}).items():
                    if field_name in candidates:
                        group_cell = remap_info.get("group_cell", "")
                        # Extract row number from cell address like "D387"
                        match = re.match(r"[A-Z]+(\d+)", group_cell.upper())
                        if match:
                            candidates[field_name].add(int(match.group(1)))
        except (json.JSONDecodeError, OSError) as e:
            logger.warning(
                "field_remaps_load_error_variant_detector",
                path=str(remaps_path),
                error=str(e),
            )

    return {field: sorted(rows) for field, rows in candidates.items()}


# Module-level cache (populated on first use)
_candidate_rows_cache: dict[str, list[int]] | None = None


def _get_candidate_rows(data_dir: Path | None = None) -> dict[str, list[int]]:
    """Get or build the candidate rows cache."""
    global _candidate_rows_cache
    if _candidate_rows_cache is None:
        _candidate_rows_cache = _load_known_candidate_rows(data_dir)
    return _candidate_rows_cache


def reset_candidate_rows_cache() -> None:
    """Reset the cached candidate rows (useful for testing)."""
    global _candidate_rows_cache
    _candidate_rows_cache = None


def _read_cell_xlsx(
    workbook: Any, sheet_name: str, col_letter: str, row: int
) -> str | None:
    """Read a single cell value from an openpyxl workbook.  Returns string or None."""
    if sheet_name not in workbook.sheetnames:
        return None
    sheet = workbook[sheet_name]
    cell_addr = f"{col_letter}{row}"
    try:
        val = sheet[cell_addr].value
        return str(val).strip() if val is not None else None
    except Exception:
        return None


def _read_cell_xlsb(
    workbook: Any, sheet_name: str, col_letter: str, row: int
) -> str | None:
    """Read a single cell value from a pyxlsb workbook.  Returns string or None."""
    if sheet_name not in workbook.sheets:
        return None

    # Convert column letter to 0-based index
    col_idx = 0
    for ch in col_letter.upper():
        col_idx = col_idx * 26 + (ord(ch) - ord("A") + 1)
    col_idx -= 1  # 0-based

    target_row = row - 1  # pyxlsb uses 0-based rows

    try:
        with workbook.get_sheet(sheet_name) as sheet:
            for ws_row in sheet.rows():
                for cell in ws_row:
                    if cell.r == target_row and cell.c == col_idx:
                        val = cell.v
                        return str(val).strip() if val is not None else None
                    # Optimization: if we've passed the target row, stop
                    if cell.r > target_row:
                        return None
    except Exception:
        return None

    return None


def detect_variant(
    file_path: str,
    file_content: bytes | None = None,
    data_dir: Path | None = None,
) -> VariantDetectionResult:
    """
    Detect whether a UW model file uses a non-production template variant.

    Opens the workbook, probes the "Assumptions (Summary)" sheet for
    label positions of Vacancy Loss, Purchase Price, and Loan Amount.
    If any field's label is NOT at the production row but IS at a known
    candidate row, the file is classified as a variant and the correct
    cell addresses are returned.

    Args:
        file_path: Path to the Excel file.
        file_content: Optional file bytes (avoids disk read).
        data_dir: Optional data directory for field_remaps.json lookup.

    Returns:
        VariantDetectionResult with is_variant flag and list of remaps.
    """
    candidate_rows = _get_candidate_rows(data_dir)
    remaps: list[VariantRemap] = []
    file_ext = Path(file_path).suffix.lower()

    # Load workbook
    try:
        if file_ext == ".xlsb":
            source = io.BytesIO(file_content) if file_content else file_path
            wb = pyxlsb.open_workbook(source)
            read_cell = _read_cell_xlsb
        else:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                if file_content:
                    wb = openpyxl.load_workbook(
                        io.BytesIO(file_content), data_only=True, read_only=True
                    )
                else:
                    wb = openpyxl.load_workbook(
                        file_path, data_only=True, read_only=True
                    )
            read_cell = _read_cell_xlsx
    except Exception as e:
        logger.debug(
            "variant_detection_workbook_load_failed",
            file=Path(file_path).name,
            error=str(e),
        )
        return VariantDetectionResult(is_variant=False, remaps=[])

    try:
        # First quick check: does the Assumptions (Summary) sheet exist?
        sheet_names = wb.sheets if file_ext == ".xlsb" else wb.sheetnames
        if _ASSUMPTIONS_SHEET not in sheet_names:
            return VariantDetectionResult(is_variant=False, remaps=[])

        # Step 1: Check vacancy field first (shared across all variants)
        # If vacancy is at the production row, this is a production template.
        vac_col, vac_prod_row, vac_label = _VARIANT_FIELDS[
            "VACANCY_LOSS_YEAR_1_RATE"
        ]

        # Check column A for the label at the production row
        prod_label = read_cell(wb, _ASSUMPTIONS_SHEET, "A", vac_prod_row)
        if prod_label and vac_label in prod_label.lower():
            # Production template -- no remaps needed
            logger.debug(
                "variant_detection_production",
                file=Path(file_path).name,
            )
            return VariantDetectionResult(is_variant=False, remaps=[])

        # Check candidate rows for vacancy label
        vac_candidates = candidate_rows.get("VACANCY_LOSS_YEAR_1_RATE", [])
        vac_found_row: int | None = None
        vac_found_label: str = ""
        for candidate_row in vac_candidates:
            if candidate_row == vac_prod_row:
                continue
            label = read_cell(wb, _ASSUMPTIONS_SHEET, "A", candidate_row)
            if label and vac_label in label.lower():
                vac_found_row = candidate_row
                vac_found_label = label
                break

        if vac_found_row is not None:
            remaps.append(
                VariantRemap(
                    field_name="VACANCY_LOSS_YEAR_1_RATE",
                    prod_cell=f"{vac_col}{vac_prod_row}",
                    variant_cell=f"{vac_col}{vac_found_row}",
                    detected_row=vac_found_row,
                    label_found=vac_found_label,
                )
            )

        # Step 2: Probe Purchase Price and Loan Amount
        for field_name in ("PURCHASE_PRICE", "LOAN_AMOUNT"):
            col, prod_row, label_substr = _VARIANT_FIELDS[field_name]
            field_candidates = candidate_rows.get(field_name, [])

            # Check production row first
            prod_label = read_cell(wb, _ASSUMPTIONS_SHEET, "A", prod_row)
            if prod_label and label_substr in prod_label.lower():
                # Field is at production position, no remap needed
                continue

            # Scan candidate rows
            for candidate_row in field_candidates:
                if candidate_row == prod_row:
                    continue
                label = read_cell(wb, _ASSUMPTIONS_SHEET, "A", candidate_row)
                if label and label_substr in label.lower():
                    remaps.append(
                        VariantRemap(
                            field_name=field_name,
                            prod_cell=f"{col}{prod_row}",
                            variant_cell=f"{col}{candidate_row}",
                            detected_row=candidate_row,
                            label_found=label,
                        )
                    )
                    break

    finally:
        if hasattr(wb, "close"):
            wb.close()

    is_variant = len(remaps) > 0

    if is_variant:
        logger.info(
            "variant_template_detected",
            file=Path(file_path).name,
            remap_count=len(remaps),
            fields=[r.field_name for r in remaps],
        )

    return VariantDetectionResult(is_variant=is_variant, remaps=remaps)


def apply_variant_remaps(
    cell_mappings: dict[str, Any],
    detection_result: VariantDetectionResult,
) -> tuple[dict[str, Any], list[dict[str, str]]]:
    """
    Apply detected variant remaps to a cell_mappings dict.

    Creates a deep copy to avoid mutating the caller's mappings.

    Args:
        cell_mappings: Original field_name -> CellMapping dict.
        detection_result: Result from detect_variant().

    Returns:
        Tuple of (modified_mappings, applied_remaps_log).
    """
    if not detection_result.is_variant:
        return cell_mappings, []

    modified = copy.deepcopy(cell_mappings)
    applied: list[dict[str, str]] = []

    for remap in detection_result.remaps:
        if remap.field_name in modified:
            original_cell = modified[remap.field_name].cell_address
            modified[remap.field_name].cell_address = remap.variant_cell
            applied.append(
                {
                    "field_name": remap.field_name,
                    "original_cell": original_cell,
                    "remapped_cell": remap.variant_cell,
                    "label_found": remap.label_found,
                }
            )

    if applied:
        logger.info(
            "variant_remaps_applied",
            remap_count=len(applied),
            fields=[a["field_name"] for a in applied],
        )

    return modified, applied


# ---------------------------------------------------------------------------
# Unit Matrix dynamic total row resolution
# ---------------------------------------------------------------------------

_UNIT_MATRIX_SHEET = "Assumptions (Unit Matrix)"

# Fields on the Unit Matrix whose row varies per file because the number
# of unit rows in the matrix depends on the property's unit count.
# Maps field_name -> (data_column, prod_row).
_UNIT_MATRIX_FIELDS: dict[str, tuple[str, int]] = {
    "TOTAL_UNITS": ("E", 548),
    "AVERAGE_UNIT_SF": ("G", 548),
}

# The label that marks the Total/Count/Average row in column C.
_TOTAL_LABEL_SUBSTR = "total/count"

# The Unit Matrix has two Total/Count/Average rows:
#   1st = end of the detailed unit list
#   2nd = end of the Unit Type Summary (Studio, 1BR, 2BR, ... 5BR)
# The production mapping (E548, G548) targets the 2nd (summary) total.
_TARGET_TOTAL_OCCURRENCE = 2


def _find_unit_matrix_total_row_xlsb(
    workbook: Any,
) -> int | None:
    """Scan the Unit Matrix sheet in a pyxlsb workbook for the summary total row.

    Returns the 1-based row number of the 2nd 'Total/Count/Average' label
    in column C, or None if the sheet is missing or the label is not found.
    """
    if _UNIT_MATRIX_SHEET not in workbook.sheets:
        return None

    try:
        occurrence = 0
        with workbook.get_sheet(_UNIT_MATRIX_SHEET) as sheet:
            for ws_row in sheet.rows():
                # Column C = index 2
                c_cell = None
                for cell in ws_row:
                    if cell.c == 2:
                        c_cell = cell
                        break
                if c_cell is None or c_cell.v is None:
                    continue
                label = str(c_cell.v).strip().lower()
                if _TOTAL_LABEL_SUBSTR in label:
                    occurrence += 1
                    if occurrence == _TARGET_TOTAL_OCCURRENCE:
                        return c_cell.r + 1  # convert 0-based to 1-based
        # If only one occurrence found, use it (some templates may differ)
        if occurrence == 1:
            # Re-scan to get the row of the single occurrence
            with workbook.get_sheet(_UNIT_MATRIX_SHEET) as sheet:
                for ws_row in sheet.rows():
                    for cell in ws_row:
                        if cell.c == 2:
                            if cell.v is not None and _TOTAL_LABEL_SUBSTR in str(cell.v).strip().lower():
                                return cell.r + 1
                            break
    except Exception:
        return None

    return None


def _find_unit_matrix_total_row_xlsx(
    workbook: Any,
) -> int | None:
    """Scan the Unit Matrix sheet in an openpyxl workbook for the summary total row.

    Returns the 1-based row number of the 2nd 'Total/Count/Average' label
    in column C, or None if the sheet is missing or the label is not found.
    """
    if _UNIT_MATRIX_SHEET not in workbook.sheetnames:
        return None

    try:
        sheet = workbook[_UNIT_MATRIX_SHEET]
        occurrence = 0
        last_match_row: int | None = None
        for row in sheet.iter_rows(min_col=3, max_col=3, values_only=False):
            cell = row[0]
            if cell.value is None:
                continue
            label = str(cell.value).strip().lower()
            if _TOTAL_LABEL_SUBSTR in label:
                occurrence += 1
                last_match_row = cell.row
                if occurrence == _TARGET_TOTAL_OCCURRENCE:
                    return cell.row
        # Fallback: if only one occurrence, use it
        if occurrence == 1 and last_match_row is not None:
            return last_match_row
    except Exception:
        return None

    return None


def resolve_unit_matrix_totals(
    file_path: str,
    file_content: bytes | None = None,
) -> list[VariantRemap]:
    """Dynamically find the correct cell addresses for Unit Matrix total fields.

    The "Assumptions (Unit Matrix)" sheet has a variable number of unit rows
    depending on the property's unit count.  The "Total/Count/Average" summary
    row (where TOTAL_UNITS and AVERAGE_UNIT_SF live) shifts accordingly.
    The production template has it at row 548 (530-unit capacity), but
    templates with fewer or more unit rows place it elsewhere.

    This function opens the workbook, scans column C for the 2nd
    "Total/Count/Average" label, and returns VariantRemap objects for any
    fields whose row differs from the production row (548).

    Args:
        file_path: Path to the Excel file.
        file_content: Optional file bytes (avoids disk read).

    Returns:
        List of VariantRemap objects (empty if row 548 is correct or
        the sheet is missing).
    """
    file_ext = Path(file_path).suffix.lower()

    try:
        if file_ext == ".xlsb":
            source = io.BytesIO(file_content) if file_content else file_path
            wb = pyxlsb.open_workbook(source)
            find_row = _find_unit_matrix_total_row_xlsb
        else:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                if file_content:
                    wb = openpyxl.load_workbook(
                        io.BytesIO(file_content), data_only=True, read_only=True
                    )
                else:
                    wb = openpyxl.load_workbook(
                        file_path, data_only=True, read_only=True
                    )
            find_row = _find_unit_matrix_total_row_xlsx
    except Exception as e:
        logger.debug(
            "unit_matrix_workbook_load_failed",
            file=Path(file_path).name,
            error=str(e),
        )
        return []

    remaps: list[VariantRemap] = []
    try:
        total_row = find_row(wb)
        if total_row is None:
            logger.debug(
                "unit_matrix_total_row_not_found",
                file=Path(file_path).name,
            )
            return []

        prod_row = 548  # production template summary total row
        if total_row == prod_row:
            return []

        for field_name, (col, _prod_row) in _UNIT_MATRIX_FIELDS.items():
            remaps.append(
                VariantRemap(
                    field_name=field_name,
                    prod_cell=f"{col}{prod_row}",
                    variant_cell=f"{col}{total_row}",
                    detected_row=total_row,
                    label_found="Total/Count/Average",
                )
            )

        if remaps:
            logger.info(
                "unit_matrix_total_row_resolved",
                file=Path(file_path).name,
                prod_row=prod_row,
                resolved_row=total_row,
                fields=[r.field_name for r in remaps],
            )

    finally:
        if hasattr(wb, "close"):
            wb.close()

    return remaps
