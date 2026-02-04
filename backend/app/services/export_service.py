"""
Excel export service with professional formatting.

Provides functionality to export data to formatted Excel files including:
- Property data exports with conditional formatting
- Deal pipeline exports with stage-based coloring
- Analytics reports with charts
- Multi-sheet workbooks
"""

from datetime import datetime
from io import BytesIO
from typing import Any

from loguru import logger

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import (
        Alignment,
        Border,
        Font,
        NamedStyle,
        PatternFill,
        Side,
    )
    from openpyxl.utils import get_column_letter

    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed - Excel export will be limited")


class ExcelExportService:
    """Service for generating formatted Excel exports."""

    # Color palette for B&R Capital branding
    COLORS = {
        "primary": "1E3A5F",  # Dark blue
        "secondary": "2E5090",  # Medium blue
        "accent": "4A90D9",  # Light blue
        "success": "28A745",  # Green
        "warning": "FFC107",  # Yellow
        "danger": "DC3545",  # Red
        "header_bg": "1E3A5F",  # Header background
        "header_fg": "FFFFFF",  # Header text
        "alt_row": "F8F9FA",  # Alternating row background
    }

    # Stage colors for deal pipeline
    STAGE_COLORS = {
        "dead": "DC3545",
        "initial_review": "BBDEFB",
        "active_review": "90CAF9",
        "under_contract": "2196F3",
        "closed": "28A745",
        "realized": "1B5E20",
    }

    def __init__(self):
        """Initialize the Excel export service."""
        self._styles_created = False
        self._header_style: NamedStyle | None = None
        self._currency_style: NamedStyle | None = None
        self._percent_style: NamedStyle | None = None
        self._date_style: NamedStyle | None = None

    def _create_styles(self, workbook: "Workbook") -> None:
        """Create reusable styles for the workbook."""
        if self._styles_created:
            return

        # Header style
        self._header_style = NamedStyle(name="header_style")
        self._header_style.font = Font(
            bold=True, color=self.COLORS["header_fg"], size=11
        )
        self._header_style.fill = PatternFill(
            start_color=self.COLORS["header_bg"],
            end_color=self.COLORS["header_bg"],
            fill_type="solid",
        )
        self._header_style.alignment = Alignment(horizontal="center", vertical="center")
        self._header_style.border = Border(
            bottom=Side(style="thin", color="000000"),
        )

        # Currency style
        self._currency_style = NamedStyle(name="currency_style")
        self._currency_style.number_format = '"$"#,##0.00'
        self._currency_style.alignment = Alignment(horizontal="right")

        # Percent style
        self._percent_style = NamedStyle(name="percent_style")
        self._percent_style.number_format = "0.00%"
        self._percent_style.alignment = Alignment(horizontal="right")

        # Date style
        self._date_style = NamedStyle(name="date_style")
        self._date_style.number_format = "YYYY-MM-DD"
        self._date_style.alignment = Alignment(horizontal="center")

        # Add styles to workbook
        try:
            workbook.add_named_style(self._header_style)
            workbook.add_named_style(self._currency_style)
            workbook.add_named_style(self._percent_style)
            workbook.add_named_style(self._date_style)
        except ValueError:
            # Styles already exist
            pass

        self._styles_created = True

    def _apply_header_style(self, cell) -> None:
        """Apply header style to a cell."""
        cell.font = Font(bold=True, color=self.COLORS["header_fg"], size=11)
        cell.fill = PatternFill(
            start_color=self.COLORS["header_bg"],
            end_color=self.COLORS["header_bg"],
            fill_type="solid",
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _auto_width_columns(self, worksheet) -> None:
        """Auto-adjust column widths based on content."""
        for column_cells in worksheet.columns:
            max_length = 0
            column = column_cells[0].column_letter
            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    max_length = max(max_length, cell_length)
                except (TypeError, AttributeError):
                    pass
            # Set width with some padding
            adjusted_width = min(max_length + 2, 50)
            worksheet.column_dimensions[column].width = adjusted_width

    def _add_alternating_rows(self, worksheet, start_row: int, end_row: int) -> None:
        """Add alternating row colors for better readability."""
        alt_fill = PatternFill(
            start_color=self.COLORS["alt_row"],
            end_color=self.COLORS["alt_row"],
            fill_type="solid",
        )
        for row_num in range(start_row, end_row + 1):
            if row_num % 2 == 0:
                for cell in worksheet[row_num]:
                    if cell.fill.patternType is None:
                        cell.fill = alt_fill

    def export_properties(
        self,
        properties: list[dict[str, Any]],
        include_analytics: bool = True,
    ) -> BytesIO:
        """
        Export properties to a formatted Excel file.

        Args:
            properties: List of property dictionaries
            include_analytics: Whether to include analytics summary sheet

        Returns:
            BytesIO buffer containing the Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")

        workbook = Workbook()
        self._create_styles(workbook)

        # Main properties sheet
        ws = workbook.active
        ws.title = "Properties"

        # Define columns
        columns = [
            ("ID", "id", None),
            ("Name", "name", None),
            ("Type", "property_type", None),
            ("Address", "address", None),
            ("City", "city", None),
            ("State", "state", None),
            ("Market", "market", None),
            ("Units/SF", "total_units", None),
            ("Year Built", "year_built", None),
            ("Occupancy %", "occupancy_rate", "percent"),
            ("Avg Rent", "avg_rent_per_unit", "currency"),
            ("NOI", "noi", "currency"),
            ("Cap Rate", "cap_rate", "percent_decimal"),
        ]

        # Write headers
        for col_num, (header, _, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            self._apply_header_style(cell)

        # Write data
        for row_num, prop in enumerate(properties, 2):
            for col_num, (_, field, fmt) in enumerate(columns, 1):
                value = prop.get(field, "")
                cell = ws.cell(row=row_num, column=col_num, value=value)

                # Apply formatting
                if fmt == "currency" and value:
                    cell.number_format = '"$"#,##0.00'
                elif fmt == "percent" and value:
                    cell.number_format = "0.0%"
                    cell.value = value / 100 if value > 1 else value
                elif fmt == "percent_decimal" and value:
                    cell.number_format = "0.00%"
                    cell.value = value / 100 if value > 1 else value

        # Apply formatting
        self._auto_width_columns(ws)
        self._add_alternating_rows(ws, 2, len(properties) + 1)

        # Add conditional formatting for occupancy
        if properties:
            occ_col = get_column_letter(10)  # Occupancy column
            ws.conditional_formatting.add(
                f"{occ_col}2:{occ_col}{len(properties) + 1}",
                ColorScaleRule(
                    start_type="num",
                    start_value=0.8,
                    start_color="F8D7DA",
                    mid_type="num",
                    mid_value=0.9,
                    mid_color="FFF3CD",
                    end_type="num",
                    end_value=1.0,
                    end_color="D4EDDA",
                ),
            )

        # Add analytics summary sheet if requested
        if include_analytics and properties:
            self._add_property_analytics_sheet(workbook, properties)

        # Save to buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        logger.info(f"Exported {len(properties)} properties to Excel")
        return buffer

    def _add_property_analytics_sheet(
        self,
        workbook: "Workbook",
        properties: list[dict[str, Any]],
    ) -> None:
        """Add analytics summary sheet to properties export."""
        ws = workbook.create_sheet(title="Analytics Summary")

        # Summary metrics
        total_properties = len(properties)
        total_units = sum(p.get("total_units", 0) or 0 for p in properties)
        total_noi = sum(p.get("noi", 0) or 0 for p in properties)
        avg_occupancy = (
            sum(p.get("occupancy_rate", 0) or 0 for p in properties) / total_properties
            if total_properties > 0
            else 0
        )
        avg_cap_rate = (
            sum(p.get("cap_rate", 0) or 0 for p in properties) / total_properties
            if total_properties > 0
            else 0
        )

        # Write summary
        summary_data = [
            ("Portfolio Summary", ""),
            ("Total Properties", total_properties),
            ("Total Units", total_units),
            ("Total NOI", total_noi),
            ("Average Occupancy", avg_occupancy / 100),
            ("Average Cap Rate", avg_cap_rate / 100),
        ]

        for row_num, (label, value) in enumerate(summary_data, 1):
            ws.cell(row=row_num, column=1, value=label)
            cell = ws.cell(row=row_num, column=2, value=value)

            if "NOI" in label:
                cell.number_format = '"$"#,##0.00'
            elif "Occupancy" in label or "Cap Rate" in label:
                cell.number_format = "0.0%"

        # Bold the header
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        self._auto_width_columns(ws)

    def export_deals(
        self,
        deals: list[dict[str, Any]],
        include_pipeline: bool = True,
    ) -> BytesIO:
        """
        Export deals to a formatted Excel file.

        Args:
            deals: List of deal dictionaries
            include_pipeline: Whether to include pipeline summary sheet

        Returns:
            BytesIO buffer containing the Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")

        workbook = Workbook()
        self._create_styles(workbook)

        # Main deals sheet
        ws = workbook.active
        ws.title = "Deals"

        # Define columns
        columns = [
            ("ID", "id", None),
            ("Name", "name", None),
            ("Type", "deal_type", None),
            ("Stage", "stage", "stage"),
            ("Priority", "priority", None),
            ("Asking Price", "asking_price", "currency"),
            ("Offer Price", "offer_price", "currency"),
            ("Projected IRR", "projected_irr", "percent"),
            ("Projected CoC", "projected_coc", "percent"),
            ("Equity Multiple", "projected_equity_multiple", None),
            ("Hold Period (Yrs)", "hold_period_years", None),
            ("Source", "source", None),
            ("Created", "created_at", "date"),
            ("Updated", "updated_at", "date"),
        ]

        # Write headers
        for col_num, (header, _, _) in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            self._apply_header_style(cell)

        # Write data with stage coloring
        for row_num, deal in enumerate(deals, 2):
            for col_num, (_, field, fmt) in enumerate(columns, 1):
                value = deal.get(field, "")
                cell = ws.cell(row=row_num, column=col_num, value=value)

                # Apply formatting
                if fmt == "currency" and value:
                    cell.number_format = '"$"#,##0'
                elif fmt == "percent" and value:
                    cell.number_format = "0.0%"
                    if isinstance(value, (int, float)) and value > 1:
                        cell.value = value
                    elif isinstance(value, (int, float)):
                        cell.value = value * 100 if value < 1 else value
                elif fmt == "date" and value:
                    cell.number_format = "YYYY-MM-DD"
                elif fmt == "stage" and value:
                    # Apply stage-based coloring
                    stage_color = self.STAGE_COLORS.get(value, "FFFFFF")
                    cell.fill = PatternFill(
                        start_color=stage_color,
                        end_color=stage_color,
                        fill_type="solid",
                    )

        # Apply formatting
        self._auto_width_columns(ws)

        # Add pipeline summary if requested
        if include_pipeline and deals:
            self._add_pipeline_summary_sheet(workbook, deals)

        # Save to buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        logger.info(f"Exported {len(deals)} deals to Excel")
        return buffer

    def _add_pipeline_summary_sheet(
        self,
        workbook: "Workbook",
        deals: list[dict[str, Any]],
    ) -> None:
        """Add pipeline summary sheet with stage breakdown."""
        ws = workbook.create_sheet(title="Pipeline Summary")

        # Count deals by stage
        stage_counts = {}
        stage_values = {}
        for deal in deals:
            stage = deal.get("stage", "unknown")
            stage_counts[stage] = stage_counts.get(stage, 0) + 1
            asking_price = deal.get("asking_price", 0) or 0
            stage_values[stage] = stage_values.get(stage, 0) + asking_price

        # Write summary
        ws.cell(row=1, column=1, value="Pipeline Summary")
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)

        headers = ["Stage", "Deal Count", "Total Value"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            self._apply_header_style(cell)

        row = 4
        for stage in [
            "lead",
            "initial_review",
            "underwriting",
            "due_diligence",
            "loi_submitted",
            "under_contract",
            "closed",
            "dead",
        ]:
            if stage in stage_counts:
                ws.cell(row=row, column=1, value=stage.replace("_", " ").title())
                ws.cell(row=row, column=2, value=stage_counts[stage])
                value_cell = ws.cell(
                    row=row, column=3, value=stage_values.get(stage, 0)
                )
                value_cell.number_format = '"$"#,##0'

                # Apply stage color
                stage_color = self.STAGE_COLORS.get(stage, "FFFFFF")
                ws.cell(row=row, column=1).fill = PatternFill(
                    start_color=stage_color,
                    end_color=stage_color,
                    fill_type="solid",
                )
                row += 1

        # Add totals
        ws.cell(row=row + 1, column=1, value="Total")
        ws.cell(row=row + 1, column=1).font = Font(bold=True)
        ws.cell(row=row + 1, column=2, value=len(deals))
        total_value_cell = ws.cell(
            row=row + 1,
            column=3,
            value=sum(stage_values.values()),
        )
        total_value_cell.number_format = '"$"#,##0'
        total_value_cell.font = Font(bold=True)

        self._auto_width_columns(ws)

    def export_analytics_report(
        self,
        dashboard_metrics: dict[str, Any],
        portfolio_analytics: dict[str, Any],
        deal_pipeline: dict[str, Any],
    ) -> BytesIO:
        """
        Export a comprehensive analytics report.

        Args:
            dashboard_metrics: Dashboard metrics data
            portfolio_analytics: Portfolio analytics data
            deal_pipeline: Deal pipeline analytics data

        Returns:
            BytesIO buffer containing the Excel file
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")

        workbook = Workbook()
        self._create_styles(workbook)

        # Dashboard Summary Sheet
        ws = workbook.active
        ws.title = "Dashboard Summary"
        self._write_dashboard_summary(ws, dashboard_metrics)

        # Portfolio Performance Sheet
        ws_portfolio = workbook.create_sheet(title="Portfolio Performance")
        self._write_portfolio_performance(ws_portfolio, portfolio_analytics)

        # Deal Pipeline Sheet
        ws_pipeline = workbook.create_sheet(title="Deal Pipeline")
        self._write_deal_pipeline(ws_pipeline, deal_pipeline)

        # Save to buffer
        buffer = BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        logger.info("Exported analytics report to Excel")
        return buffer

    def _write_dashboard_summary(
        self,
        ws,
        metrics: dict[str, Any],
    ) -> None:
        """Write dashboard summary to worksheet."""
        ws.cell(row=1, column=1, value="Dashboard Summary")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)
        ws.cell(
            row=2,
            column=1,
            value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        )

        # Portfolio Summary Section
        row = 4
        ws.cell(row=row, column=1, value="Portfolio Summary")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        portfolio = metrics.get("portfolio_summary", {})
        summary_items = [
            ("Total Properties", portfolio.get("total_properties", 0)),
            ("Total Units", portfolio.get("total_units", 0)),
            ("Total SF", portfolio.get("total_sf", 0)),
            ("Total Value", portfolio.get("total_value", 0)),
            ("Avg Occupancy", portfolio.get("avg_occupancy", 0)),
            ("Avg Cap Rate", portfolio.get("avg_cap_rate", 0)),
        ]

        for label, value in summary_items:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=value)
            if "Value" in label:
                cell.number_format = '"$"#,##0'
            elif "Occupancy" in label or "Cap Rate" in label:
                cell.number_format = '0.0"%"'
            row += 1

        # KPIs Section
        row += 1
        ws.cell(row=row, column=1, value="Key Performance Indicators")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        kpis = metrics.get("kpis", {})
        kpi_items = [
            ("YTD NOI Growth", kpis.get("ytd_noi_growth", 0)),
            ("YTD Rent Growth", kpis.get("ytd_rent_growth", 0)),
            ("Deals in Pipeline", kpis.get("deals_in_pipeline", 0)),
            ("Deals Closed YTD", kpis.get("deals_closed_ytd", 0)),
            ("Capital Deployed YTD", kpis.get("capital_deployed_ytd", 0)),
        ]

        for label, value in kpi_items:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=value)
            if "Growth" in label:
                cell.number_format = '0.0"%"'
            elif "Capital" in label:
                cell.number_format = '"$"#,##0'
            row += 1

        self._auto_width_columns(ws)

    def _write_portfolio_performance(
        self,
        ws,
        analytics: dict[str, Any],
    ) -> None:
        """Write portfolio performance to worksheet."""
        ws.cell(row=1, column=1, value="Portfolio Performance")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)

        # Performance metrics
        row = 3
        ws.cell(row=row, column=1, value="Performance Metrics")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        perf = analytics.get("performance", {})
        perf_items = [
            ("Total Return", perf.get("total_return", 0)),
            ("Income Return", perf.get("income_return", 0)),
            ("Appreciation Return", perf.get("appreciation_return", 0)),
            ("Benchmark Return", perf.get("benchmark_return", 0)),
            ("Alpha", perf.get("alpha", 0)),
        ]

        for label, value in perf_items:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=value)
            cell.number_format = '0.0"%"'
            row += 1

        self._auto_width_columns(ws)

    def _write_deal_pipeline(
        self,
        ws,
        pipeline: dict[str, Any],
    ) -> None:
        """Write deal pipeline analytics to worksheet."""
        ws.cell(row=1, column=1, value="Deal Pipeline Analytics")
        ws.cell(row=1, column=1).font = Font(bold=True, size=16)

        # Funnel metrics
        row = 3
        ws.cell(row=row, column=1, value="Pipeline Funnel")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        headers = ["Stage", "Count"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            self._apply_header_style(cell)
        row += 1

        funnel = pipeline.get("funnel", {})
        for stage, count in funnel.items():
            ws.cell(row=row, column=1, value=stage.replace("_", " ").title())
            ws.cell(row=row, column=2, value=count)

            stage_color = self.STAGE_COLORS.get(stage, "FFFFFF")
            ws.cell(row=row, column=1).fill = PatternFill(
                start_color=stage_color,
                end_color=stage_color,
                fill_type="solid",
            )
            row += 1

        # Conversion rates
        row += 1
        ws.cell(row=row, column=1, value="Conversion Rates")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        rates = pipeline.get("conversion_rates", {})
        for metric, rate in rates.items():
            ws.cell(row=row, column=1, value=metric.replace("_", " ").title())
            cell = ws.cell(row=row, column=2, value=rate)
            cell.number_format = '0.0"%"'
            row += 1

        self._auto_width_columns(ws)


# Service singleton
_excel_service: ExcelExportService | None = None


def get_excel_service() -> ExcelExportService:
    """Get the Excel export service singleton."""
    global _excel_service
    if _excel_service is None:
        _excel_service = ExcelExportService()
    return _excel_service
