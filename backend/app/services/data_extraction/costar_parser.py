"""
CoStar Excel Parser -- transforms wide-format CoStar exports to normalized DB rows.

Handles both MSA-level and Submarket-level exports.
Column layout (both files, sheet "DataExport"):
  A (0): Property Class Name  (always "Multi-Family")
  B (1): Slice               (always "All")
  C (2): As Of               (e.g. "2026 Q1")
  D (3): Geography Name      (e.g. "Phoenix - AZ USA" or "Phoenix - AZ USA - Tempe")
  E (4): Geography Code
  F (5): Property Type
  G (6): Forecast Scenario
  H (7): CBSA Code
  I (8): Geography Type      ("Metro" or "Submarket")
  J (9): Concept Name        (63 concepts)
  K+ (10+): Quarterly data columns ("1982 Q1" .. "2031 Q1" for MSA; "2000 Q1" .. "2031 Q1" for submarket)

Usage:
  python -m app.services.data_extraction.costar_parser
"""

from __future__ import annotations

import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any

import asyncpg
import openpyxl
import structlog

logger = structlog.get_logger(__name__)

# Quarter string -> month mapping
QUARTER_TO_MONTH: dict[str, int] = {"Q1": 1, "Q2": 4, "Q3": 7, "Q4": 10}

# Metadata column indices
_COL_GEO_NAME = 3
_COL_GEO_CODE = 4
_COL_GEO_TYPE = 8
_COL_CONCEPT = 9
_FIRST_DATA_COL = 10

# Default batch size for upserts
_BATCH_SIZE = 1000


class CoStarParser:
    """Parse CoStar wide-format Excel exports and upsert into PostgreSQL.

    Each data row in the Excel has a single concept for a single geography,
    with quarterly values spread across columns.  This parser transposes
    that into normalized rows:
        (geography_type, geography_name, concept, date, value, is_forecast, source_file)
    """

    def __init__(self, db_url: str, data_dir: str) -> None:
        """Initialize with database URL and CoStar data directory path.

        Args:
            db_url: PostgreSQL connection string
                (e.g. ``postgresql://user:pass@localhost:5432/market_analysis``).
            data_dir: Filesystem path to directory containing CoStar ``.xlsx`` exports.
        """
        self._db_url = db_url
        self._data_dir = Path(data_dir)
        self._pool: asyncpg.Pool | None = None
        self._current_quarter_start = self._compute_current_quarter_start()

    # ------------------------------------------------------------------
    # Public API
    # ------------------------------------------------------------------

    async def parse_all(self) -> dict[str, Any]:
        """Parse all CoStar Excel files and upsert into database.

        Returns:
            Summary dict with keys ``status``, ``files_processed``,
            ``records_upserted``, and ``errors``.
        """
        if not self._data_dir.exists():
            msg = f"CoStar data directory not found: {self._data_dir}"
            logger.error(msg)
            return {"status": "error", "message": msg, "records_upserted": 0}

        self._pool = await asyncpg.create_pool(self._db_url, min_size=1, max_size=4)

        log_id = await self._start_extraction_log()
        total_records = 0
        files_processed = 0
        errors: list[str] = []

        # Discover xlsx files, skip the submarket *list* file and tiny Zone.Identifier stubs
        xlsx_files = sorted(
            f
            for f in self._data_dir.glob("*.xlsx")
            if f.stat().st_size > 1000 and "Submarket List" not in f.name
        )

        for filepath in xlsx_files:
            try:
                geo_type = self._detect_geography_type(filepath)
                if geo_type == "Metro":
                    count = await self.parse_msa_file(filepath)
                else:
                    count = await self.parse_submarket_file(filepath)
                total_records += count
                files_processed += 1
            except Exception as exc:
                logger.error("costar_parse_error", file=filepath.name, error=str(exc))
                errors.append(f"{filepath.name}: {exc}")

        # Refresh materialized view if it exists
        await self._try_refresh_materialized_view()

        status = "success" if not errors else "partial_error"
        await self._finish_extraction_log(
            log_id, status, total_records, errors, files_processed
        )
        await self._pool.close()
        self._pool = None

        summary: dict[str, Any] = {
            "status": status,
            "files_processed": files_processed,
            "records_upserted": total_records,
            "errors": errors,
        }
        logger.info("costar_extraction_complete", **summary)
        return summary

    async def parse_msa_file(self, filepath: Path) -> int:
        """Parse the MSA-level market data export.

        Args:
            filepath: Path to the MSA Excel file.

        Returns:
            Number of records upserted.
        """
        return await self._parse_file(filepath)

    async def parse_submarket_file(self, filepath: Path) -> int:
        """Parse the submarket-level data export.

        Args:
            filepath: Path to the submarket Excel file.

        Returns:
            Number of records upserted.
        """
        return await self._parse_file(filepath)

    def _parse_quarter_header(self, header: str) -> date | None:
        """Parse quarter column header like ``'2025 Q1'`` to a :class:`date`.

        Mapping: Q1 -> Jan 1, Q2 -> Apr 1, Q3 -> Jul 1, Q4 -> Oct 1.

        Returns:
            ``None`` if the header does not match the expected pattern.
        """
        match = re.match(r"(\d{4})\s+(Q[1-4])", str(header).strip())
        if not match:
            return None
        year = int(match.group(1))
        month = QUARTER_TO_MONTH[match.group(2)]
        return date(year, month, 1)

    def _is_forecast(self, dt: date) -> bool:
        """Return ``True`` if *dt* is in the future (forecast data)."""
        return dt > self._current_quarter_start

    # ------------------------------------------------------------------
    # Internal helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _compute_current_quarter_start() -> date:
        """Return the first day of the current calendar quarter."""
        now = datetime.now()
        quarter_month = ((now.month - 1) // 3) * 3 + 1
        return date(now.year, quarter_month, 1)

    @staticmethod
    def _detect_geography_type(filepath: Path) -> str:
        """Heuristic: if filename contains 'Submarket' treat as submarket data."""
        return "Submarket" if "Submarket" in filepath.name else "Metro"

    async def _parse_file(self, filepath: Path) -> int:
        """Core parsing logic shared by MSA and submarket files."""
        log = logger.bind(file=filepath.name)
        log.info("costar_parse_start")

        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        try:
            ws = wb["DataExport"]
        except KeyError:
            ws = wb[wb.sheetnames[0]]

        # --- Build date column index from header row ---
        header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        date_columns: list[tuple[int, date, bool]] = []
        for col_idx, cell_val in enumerate(header_row):
            if col_idx < _FIRST_DATA_COL:
                continue
            dt = self._parse_quarter_header(str(cell_val) if cell_val else "")
            if dt is not None:
                date_columns.append((col_idx, dt, self._is_forecast(dt)))

        log.info("costar_columns_parsed", quarter_columns=len(date_columns))

        # --- Extract normalized records ---
        records: list[tuple[str, str, str | None, str, date, float, bool, str]] = []
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            row_count += 1
            geo_name = str(row[_COL_GEO_NAME]).strip() if row[_COL_GEO_NAME] else None
            geo_type = str(row[_COL_GEO_TYPE]).strip() if row[_COL_GEO_TYPE] else None
            geo_code = str(row[_COL_GEO_CODE]).strip() if row[_COL_GEO_CODE] else None
            concept = str(row[_COL_CONCEPT]).strip() if row[_COL_CONCEPT] else None

            if not geo_name or not concept or not geo_type:
                continue

            for col_idx, dt, is_fc in date_columns:
                val = row[col_idx] if col_idx < len(row) else None
                if val is None:
                    continue
                try:
                    float_val = float(val)
                except (ValueError, TypeError):
                    continue

                records.append(
                    (
                        geo_type,
                        geo_name,
                        geo_code,
                        concept,
                        dt,
                        float_val,
                        is_fc,
                        filepath.name,
                    )
                )

        wb.close()
        log.info("costar_records_extracted", data_rows=row_count, records=len(records))

        if not records:
            return 0

        upserted = await self._batch_upsert(records)
        log.info("costar_upsert_complete", upserted=upserted)
        return upserted

    async def _batch_upsert(
        self,
        records: list[tuple[str, str, str | None, str, date, float, bool, str]],
    ) -> int:
        """Upsert records into ``costar_timeseries`` in batches of ``_BATCH_SIZE``."""
        assert self._pool is not None
        upsert_sql = """
            INSERT INTO costar_timeseries
                (geography_type, geography_name, geography_code, concept, date, value, is_forecast, source_file)
            SELECT
                r.geography_type, r.geography_name, r.geography_code,
                r.concept, r.date, r.value, r.is_forecast, r.source_file
            FROM unnest(
                $1::text[], $2::text[], $3::text[], $4::text[],
                $5::date[], $6::double precision[], $7::boolean[], $8::text[]
            ) AS r(geography_type, geography_name, geography_code, concept, date, value, is_forecast, source_file)
            ON CONFLICT (geography_type, geography_name, concept, date)
            DO UPDATE SET
                value        = EXCLUDED.value,
                is_forecast  = EXCLUDED.is_forecast,
                source_file  = EXCLUDED.source_file,
                geography_code = EXCLUDED.geography_code,
                updated_at   = NOW()
        """

        total = 0
        for start in range(0, len(records), _BATCH_SIZE):
            batch = records[start : start + _BATCH_SIZE]
            # Decompose tuple list into columnar arrays for unnest
            geo_types = [r[0] for r in batch]
            geo_names = [r[1] for r in batch]
            geo_codes = [r[2] for r in batch]
            concepts = [r[3] for r in batch]
            dates = [r[4] for r in batch]
            values = [r[5] for r in batch]
            forecasts = [r[6] for r in batch]
            sources = [r[7] for r in batch]

            async with self._pool.acquire() as conn:
                await conn.execute(
                    upsert_sql,
                    geo_types,
                    geo_names,
                    geo_codes,
                    concepts,
                    dates,
                    values,
                    forecasts,
                    sources,
                )
            total += len(batch)
            logger.debug(
                "costar_batch_upserted", batch_size=len(batch), total_so_far=total
            )

        return total

    # ------------------------------------------------------------------
    # Extraction log helpers
    # ------------------------------------------------------------------

    async def _start_extraction_log(self) -> int | None:
        """Insert a running extraction_log entry and return its id."""
        assert self._pool is not None
        try:
            async with self._pool.acquire() as conn:
                row = await conn.fetchrow(
                    "INSERT INTO extraction_log (source, status) VALUES ('costar', 'running') RETURNING id"
                )
                return row["id"] if row else None
        except Exception as exc:
            logger.warning("costar_extraction_log_insert_failed", error=str(exc))
            return None

    async def _finish_extraction_log(
        self,
        log_id: int | None,
        status: str,
        records: int,
        errors: list[str],
        files_processed: int,
    ) -> None:
        """Update the extraction_log entry with final results."""
        if log_id is None or self._pool is None:
            return
        try:
            async with self._pool.acquire() as conn:
                await conn.execute(
                    """
                    UPDATE extraction_log
                    SET finished_at = NOW(),
                        status = $1,
                        records_upserted = $2,
                        error_message = $3,
                        details = $4
                    WHERE id = $5
                    """,
                    status,
                    records,
                    "; ".join(errors) if errors else None,
                    f'{{"files_processed": {files_processed}}}',
                    log_id,
                )
        except Exception as exc:
            logger.warning("costar_extraction_log_update_failed", error=str(exc))

    async def _try_refresh_materialized_view(self) -> None:
        """Refresh the ``costar_latest`` materialized view if it exists."""
        if self._pool is None:
            return
        try:
            async with self._pool.acquire() as conn:
                await conn.execute(
                    "REFRESH MATERIALIZED VIEW CONCURRENTLY costar_latest"
                )
        except Exception:
            try:
                async with self._pool.acquire() as conn:
                    await conn.execute("REFRESH MATERIALIZED VIEW costar_latest")
            except Exception as exc:
                logger.debug("costar_matview_refresh_skipped", reason=str(exc))


# ----------------------------------------------------------------------
# Synchronous convenience wrapper (used by scheduler / CLI)
# ----------------------------------------------------------------------


def run_costar_extraction_sync(
    db_url: str | None = None, data_dir: str | None = None
) -> dict:
    """Synchronous entry-point that delegates to the async parser.

    Falls back to ``settings.MARKET_ANALYSIS_DB_URL`` and ``settings.COSTAR_DATA_DIR``
    when arguments are not provided.
    """
    from app.core.config import settings

    db_url = db_url or settings.MARKET_ANALYSIS_DB_URL
    data_dir = data_dir or settings.COSTAR_DATA_DIR

    if not db_url:
        logger.error("MARKET_ANALYSIS_DB_URL not configured")
        return {
            "status": "error",
            "message": "No DB URL configured",
            "records_upserted": 0,
        }

    import asyncio

    parser = CoStarParser(db_url=db_url, data_dir=data_dir)
    return asyncio.run(parser.parse_all())


# ----------------------------------------------------------------------
# Script entry-point
# ----------------------------------------------------------------------


async def main() -> None:
    """Load settings and run ``parse_all()``."""
    from app.core.config import settings

    db_url = settings.MARKET_ANALYSIS_DB_URL
    if not db_url:
        logger.error("MARKET_ANALYSIS_DB_URL not set -- cannot run CoStar extraction")
        sys.exit(1)

    data_dir = settings.COSTAR_DATA_DIR
    parser = CoStarParser(db_url=db_url, data_dir=data_dir)
    result = await parser.parse_all()

    print(result)
    sys.exit(0 if result["status"] == "success" else 1)


if __name__ == "__main__":
    import asyncio

    asyncio.run(main())
