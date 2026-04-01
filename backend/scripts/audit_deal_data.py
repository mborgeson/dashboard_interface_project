"""
Dashboard Deal Data Accuracy & Completeness Audit

Audits every data point displayed on the dashboard for deals:
  Phase 1: DB-level completeness audit (all enrichment + proforma fields)
  Phase 2: Excel source file verification (N/A/error/missing values)
  Phase 3: Fix discrepancies (where Excel has real values but DB has N/A)
  Phase 4: Generate N/A documentation and summary report

Usage:
    cd backend/
    python scripts/audit_deal_data.py                    # Full audit, dry-run fixes
    python scripts/audit_deal_data.py --fix              # Full audit + apply fixes
    python scripts/audit_deal_data.py --deal "44 Monroe" # Audit specific deal
    python scripts/audit_deal_data.py --db-only          # Skip Excel verification
"""

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import UTC, datetime
from pathlib import Path
from typing import Any
from uuid import uuid4

from loguru import logger

# Ensure `app` is importable when running from backend/
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from sqlalchemy import and_, func, or_, select, text
from sqlalchemy.orm import Session

from app.db.session import SessionLocal

# ── Constants ─────────────────────────────────────────────────────────────────

REPORTS_DIR = Path(__file__).resolve().parent.parent / "data" / "audit_reports"
UW_MODEL_DIR = Path(__file__).resolve().parent.parent / "data" / "uw_model_files"
GROUPS_DIR = Path(__file__).resolve().parent.parent / "data" / "extraction_groups"
GROUPS_JSON = GROUPS_DIR / "groups.json"

# 33 core enrichment fields (from enrichment.py _ENRICHMENT_FIELDS)
ENRICHMENT_FIELDS = [
    "TOTAL_UNITS",
    "AVERAGE_UNIT_SF",
    "CURRENT_OWNER",
    "LAST_SALE_PRICE_PER_UNIT",
    "LAST_SALE_DATE",
    "T12_RETURN_ON_COST",
    "LP_RETURNS_IRR",
    "LP_RETURNS_MOIC",
    "UNLEVERED_RETURNS_IRR",
    "UNLEVERED_RETURNS_MOIC",
    "LEVERED_RETURNS_IRR",
    "LEVERED_RETURNS_MOIC",
    "PROPERTY_CITY",
    "SUBMARKET",
    "YEAR_BUILT",
    "YEAR_RENOVATED",
    "VACANCY_LOSS_YEAR_1_RATE",
    "BAD_DEBTS_YEAR_1_RATE",
    "OTHER_LOSS_YEAR_1_RATE",
    "CONCESSIONS_YEAR_1_RATE",
    "NET_OPERATING_INCOME_MARGIN",
    "PURCHASE_PRICE",
    "TOTAL_ACQUISITION_BUDGET",
    "BASIS_UNIT_AT_CLOSE",
    "T12_RETURN_ON_PP",
    "T3_RETURN_ON_PP",
    "LOAN_AMOUNT",
    "EQUITY_LP_CAPITAL",
    "EXIT_PERIOD_MONTHS",
    "EXIT_CAP_RATE",
    "T3_RETURN_ON_COST",
    "PROPERTY_LATITUDE",
    "PROPERTY_LONGITUDE",
]

# 27 proforma fields (from enrichment.py PROFORMA_FIELDS)
PROFORMA_FIELDS = [
    "LEVERED_RETURNS_IRR_YR2",
    "LEVERED_RETURNS_IRR_YR3",
    "LEVERED_RETURNS_IRR_YR7",
    "LEVERED_RETURNS_MOIC_YR2",
    "LEVERED_RETURNS_MOIC_YR3",
    "LEVERED_RETURNS_MOIC_YR7",
    "UNLEVERED_RETURNS_IRR_YR2",
    "UNLEVERED_RETURNS_IRR_YR3",
    "UNLEVERED_RETURNS_IRR_YR7",
    "UNLEVERED_RETURNS_MOIC_YR2",
    "UNLEVERED_RETURNS_MOIC_YR3",
    "UNLEVERED_RETURNS_MOIC_YR7",
    "NOI_PER_UNIT_YR2",
    "NOI_PER_UNIT_YR3",
    "NOI_PER_UNIT_YR5",
    "NOI_PER_UNIT_YR7",
    "CAP_RATE_ALL_IN_YR3",
    "CAP_RATE_ALL_IN_YR5",
    "COC_YR5",
    "DSCR_T3",
    "DSCR_YR5",
    "PROFORMA_NOI_YR1",
    "PROFORMA_NOI_YR2",
    "PROFORMA_NOI_YR3",
    "PROFORMA_DSCR_YR1",
    "PROFORMA_DSCR_YR2",
    "PROFORMA_DSCR_YR3",
    "PROFORMA_DEBT_YIELD_YR1",
    "PROFORMA_DEBT_YIELD_YR2",
    "PROFORMA_DEBT_YIELD_YR3",
]

ALL_AUDIT_FIELDS = ENRICHMENT_FIELDS + PROFORMA_FIELDS

# Fields with >0 guard in _apply_extraction_fields (enrichment.py)
POSITIVE_GUARD_FIELDS = {
    "T12_RETURN_ON_PP",
    "T3_RETURN_ON_PP",
    "T12_RETURN_ON_COST",
    "T3_RETURN_ON_COST",
    "BASIS_UNIT_AT_CLOSE",
}

# Template deal name patterns to exclude
TEMPLATE_PATTERNS = [
    r"^\[.*\]",  # [Investment Name], [Deal Name], [City]
    r"^Template",
    r"^Sample",
    r"^Property \(",  # "Property (Tempe, AZ)"
    r"^Deal \(",  # "Deal (Unknown, AZ)"
]

# Human-readable field descriptions for N/A documentation
FIELD_DESCRIPTIONS: dict[str, str] = {
    "TOTAL_UNITS": "Total number of apartment units",
    "AVERAGE_UNIT_SF": "Average square footage per unit",
    "CURRENT_OWNER": "Current property owner name",
    "LAST_SALE_PRICE_PER_UNIT": "Last sale price per unit",
    "LAST_SALE_DATE": "Date of last property sale",
    "T12_RETURN_ON_COST": "Trailing 12-month cap rate on total cost",
    "LP_RETURNS_IRR": "Limited Partner Internal Rate of Return",
    "LP_RETURNS_MOIC": "Limited Partner Multiple on Invested Capital",
    "UNLEVERED_RETURNS_IRR": "Unlevered Internal Rate of Return",
    "UNLEVERED_RETURNS_MOIC": "Unlevered Multiple on Invested Capital",
    "LEVERED_RETURNS_IRR": "Levered Internal Rate of Return",
    "LEVERED_RETURNS_MOIC": "Levered Multiple on Invested Capital",
    "PROPERTY_CITY": "Property city",
    "SUBMARKET": "CoStar submarket cluster",
    "YEAR_BUILT": "Year property was constructed",
    "YEAR_RENOVATED": "Year of last renovation",
    "VACANCY_LOSS_YEAR_1_RATE": "Year 1 vacancy loss rate",
    "BAD_DEBTS_YEAR_1_RATE": "Year 1 bad debt rate",
    "OTHER_LOSS_YEAR_1_RATE": "Year 1 other loss rate",
    "CONCESSIONS_YEAR_1_RATE": "Year 1 concessions rate",
    "NET_OPERATING_INCOME_MARGIN": "NOI as a percentage of EGI",
    "PURCHASE_PRICE": "Total purchase price",
    "TOTAL_ACQUISITION_BUDGET": "Total acquisition budget (purchase + reno + closing)",
    "BASIS_UNIT_AT_CLOSE": "Total going-in basis per unit at close",
    "T12_RETURN_ON_PP": "Trailing 12-month cap rate on purchase price",
    "T3_RETURN_ON_PP": "Trailing 3-month cap rate on purchase price",
    "LOAN_AMOUNT": "Total loan amount",
    "EQUITY_LP_CAPITAL": "LP equity capital contribution",
    "EXIT_PERIOD_MONTHS": "Projected hold period in months",
    "EXIT_CAP_RATE": "Projected exit cap rate",
    "T3_RETURN_ON_COST": "Trailing 3-month cap rate on total cost",
    "PROPERTY_LATITUDE": "Property latitude coordinate",
    "PROPERTY_LONGITUDE": "Property longitude coordinate",
    "LEVERED_RETURNS_IRR_YR2": "Levered IRR at Year 2",
    "LEVERED_RETURNS_IRR_YR3": "Levered IRR at Year 3",
    "LEVERED_RETURNS_IRR_YR7": "Levered IRR at Year 7",
    "LEVERED_RETURNS_MOIC_YR2": "Levered MOIC at Year 2",
    "LEVERED_RETURNS_MOIC_YR3": "Levered MOIC at Year 3",
    "LEVERED_RETURNS_MOIC_YR7": "Levered MOIC at Year 7",
    "UNLEVERED_RETURNS_IRR_YR2": "Unlevered IRR at Year 2",
    "UNLEVERED_RETURNS_IRR_YR3": "Unlevered IRR at Year 3",
    "UNLEVERED_RETURNS_IRR_YR7": "Unlevered IRR at Year 7",
    "UNLEVERED_RETURNS_MOIC_YR2": "Unlevered MOIC at Year 2",
    "UNLEVERED_RETURNS_MOIC_YR3": "Unlevered MOIC at Year 3",
    "UNLEVERED_RETURNS_MOIC_YR7": "Unlevered MOIC at Year 7",
    "NOI_PER_UNIT_YR2": "NOI per unit at Year 2",
    "NOI_PER_UNIT_YR3": "NOI per unit at Year 3",
    "NOI_PER_UNIT_YR5": "NOI per unit at Year 5",
    "NOI_PER_UNIT_YR7": "NOI per unit at Year 7",
    "CAP_RATE_ALL_IN_YR3": "All-in cap rate at Year 3",
    "CAP_RATE_ALL_IN_YR5": "All-in cap rate at Year 5",
    "COC_YR5": "Cash-on-cash return at Year 5",
    "DSCR_T3": "Debt Service Coverage Ratio (trailing 3-month)",
    "DSCR_YR5": "Debt Service Coverage Ratio at Year 5",
    "PROFORMA_NOI_YR1": "Proforma NOI Year 1",
    "PROFORMA_NOI_YR2": "Proforma NOI Year 2",
    "PROFORMA_NOI_YR3": "Proforma NOI Year 3",
    "PROFORMA_DSCR_YR1": "Proforma DSCR Year 1",
    "PROFORMA_DSCR_YR2": "Proforma DSCR Year 2",
    "PROFORMA_DSCR_YR3": "Proforma DSCR Year 3",
    "PROFORMA_DEBT_YIELD_YR1": "Proforma Debt Yield Year 1",
    "PROFORMA_DEBT_YIELD_YR2": "Proforma Debt Yield Year 2",
    "PROFORMA_DEBT_YIELD_YR3": "Proforma Debt Yield Year 3",
}


# ── Data classes ──────────────────────────────────────────────────────────────


@dataclass
class DealInfo:
    deal_id: int
    deal_name: str
    property_id: int
    property_name: str
    stage: str


@dataclass
class FieldAudit:
    deal_id: int
    deal_name: str
    property_id: int
    field_name: str
    status: str  # has_value, n_a_placeholder, error, missing, guarded_zero
    value_text: str | None = None
    value_numeric: float | None = None
    is_error: bool = False
    error_category: str | None = None
    source_file: str | None = None
    sheet_name: str | None = None
    cell_address: str | None = None
    confidence_score: float | None = None
    dashboard_shows: str = "N/A"  # what the frontend would display


@dataclass
class CellVerification:
    deal_name: str
    field_name: str
    source_file: str
    sheet_name: str
    cell_address: str
    db_status: str
    db_value_text: str | None
    db_value_numeric: float | None
    actual_cell_value: Any = None
    actual_cell_type: str = ""  # numeric, string, formula_error, empty, n_a_text
    verdict: str = ""  # correct_na, correct_error, discrepancy, extraction_missed


# ── Helpers ───────────────────────────────────────────────────────────────────


def is_template_deal(name: str) -> bool:
    """Return True if deal name matches template/placeholder patterns."""
    for pattern in TEMPLATE_PATTERNS:
        if re.match(pattern, name, re.IGNORECASE):
            return True
    return False


def column_to_index(col_str: str) -> int:
    """Convert Excel column letters to 0-based index. A=0, B=1, ..., AA=26."""
    result = 0
    for char in col_str.upper():
        result = result * 26 + (ord(char) - ord("A") + 1)
    return result - 1


def parse_cell_address(cell_address: str) -> tuple[int, int] | None:
    """Parse 'A1' → (row_0based, col_0based). Returns None if invalid."""
    clean = cell_address.replace("$", "").upper()
    match = re.match(r"^([A-Z]+)(\d+)$", clean)
    if not match:
        return None
    col_str, row_str = match.groups()
    return (int(row_str) - 1, column_to_index(col_str))


def read_xlsb_cell(file_path: Path, sheet_name: str, cell_address: str) -> Any:
    """Read a single cell from a .xlsb file using pyxlsb with sheet caching."""
    import pyxlsb

    parsed = parse_cell_address(cell_address)
    if parsed is None:
        return None
    target_row, target_col = parsed

    try:
        with pyxlsb.open_workbook(str(file_path)) as wb:
            if sheet_name not in wb.sheets:
                return f"__SHEET_MISSING:{sheet_name}__"
            with wb.get_sheet(sheet_name) as sheet:
                for row in sheet.rows():
                    for cell in row:
                        if cell.r == target_row and cell.c == target_col:
                            return cell.v
    except Exception as e:
        return f"__FILE_ERROR:{e}__"
    return None  # Cell not found (empty)


def read_xlsb_cells_batch(
    file_path: Path, cells: list[tuple[str, str]]
) -> dict[tuple[str, str], Any]:
    """Read multiple cells from a .xlsb file efficiently using sheet caching.

    Args:
        file_path: Path to the .xlsb file
        cells: List of (sheet_name, cell_address) tuples

    Returns:
        dict mapping (sheet_name, cell_address) → cell value
    """
    import pyxlsb

    results: dict[tuple[str, str], Any] = {}
    # Group cells by sheet for efficient reading
    by_sheet: dict[str, list[tuple[str, tuple[int, int]]]] = defaultdict(list)
    for sheet_name, cell_address in cells:
        parsed = parse_cell_address(cell_address)
        if parsed is None:
            results[(sheet_name, cell_address)] = None
            continue
        by_sheet[sheet_name].append((cell_address, parsed))

    try:
        with pyxlsb.open_workbook(str(file_path)) as wb:
            for sheet_name, cell_list in by_sheet.items():
                if sheet_name not in wb.sheets:
                    for cell_addr, _ in cell_list:
                        results[(sheet_name, cell_addr)] = (
                            f"__SHEET_MISSING:{sheet_name}__"
                        )
                    continue

                # Build full sheet cache for O(1) lookups
                cell_cache: dict[tuple[int, int], Any] = {}
                with wb.get_sheet(sheet_name) as sheet:
                    for row in sheet.rows():
                        for cell in row:
                            cell_cache[(cell.r, cell.c)] = cell.v

                # Look up each requested cell
                for cell_addr, (target_row, target_col) in cell_list:
                    results[(sheet_name, cell_addr)] = cell_cache.get(
                        (target_row, target_col)
                    )
    except Exception as e:
        for sheet_name, cell_address in cells:
            if (sheet_name, cell_address) not in results:
                results[(sheet_name, cell_address)] = f"__FILE_ERROR:{e}__"

    return results


def classify_cell_value(value: Any) -> tuple[str, str]:
    """Classify an actual Excel cell value. Returns (type, display_value)."""
    if value is None:
        return "empty", ""
    if isinstance(value, str):
        v = value.strip()
        if not v:
            return "empty", ""
        if v.upper() in ("N/A", "NA", "#N/A", "TBD", "TBA", "-"):
            return "n_a_text", v
        if v.startswith("#") and v.endswith("!"):
            return "formula_error", v
        return "string", v
    if isinstance(value, (int, float)):
        import math

        if math.isnan(value) or math.isinf(value):
            return "empty", ""
        return "numeric", str(value)
    return "other", str(value)


def load_groups_json() -> dict:
    """Load groups.json and return the full data."""
    with open(GROUPS_JSON) as f:
        return json.load(f)


def find_file_group(source_file: str, groups_data: dict) -> tuple[str, dict | None]:
    """Find which group a source file belongs to. Returns (group_name, ref_mapping)."""
    file_name = Path(source_file).name
    for group in groups_data.get("groups", []):
        for gfile in group.get("files", []):
            if Path(gfile.get("file_path", "")).name == file_name:
                group_name = group["group_name"]
                ref_path = GROUPS_DIR / group_name / "reference_mapping.json"
                ref_mapping = None
                if ref_path.exists():
                    with open(ref_path) as f:
                        ref_mapping = json.load(f)
                return group_name, ref_mapping
    return "", None


def get_field_cell_from_mapping(
    ref_mapping: dict, field_name: str
) -> tuple[str, str] | None:
    """Get (sheet_name, cell_address) for a field from reference_mapping.json."""
    if not ref_mapping:
        return None
    for mapping in ref_mapping.get("mappings", []):
        if mapping.get("field_name") == field_name:
            sheet = mapping.get("production_sheet") or mapping.get("source_sheet")
            cell = mapping.get("production_cell") or mapping.get("source_cell")
            if sheet and cell:
                return sheet, cell
    return None


# ── Phase 1: DB-Level Completeness Audit ──────────────────────────────────────


def phase1_db_audit(db: Session, deal_filter: str | None = None) -> list[FieldAudit]:
    """Audit all enrichment + proforma fields for every deal in the DB.

    Returns a FieldAudit record for each (deal, field) combination.
    """
    logger.info("Phase 1: Starting DB-level completeness audit...")

    # Get all active deals with property_id
    deal_query = text("""
        SELECT d.id, d.name, d.property_id, p.name as property_name, d.stage
        FROM deals d
        JOIN properties p ON p.id = d.property_id
        WHERE d.is_deleted = false
          AND d.property_id IS NOT NULL
    """)
    rows = db.execute(deal_query).fetchall()
    deals = [
        DealInfo(
            deal_id=r[0],
            deal_name=r[1],
            property_id=r[2],
            property_name=r[3] or r[1],
            stage=r[4],
        )
        for r in rows
    ]
    logger.info(f"Found {len(deals)} active deals with property_id")

    # Filter out template deals
    real_deals = [d for d in deals if not is_template_deal(d.deal_name)]
    template_count = len(deals) - len(real_deals)
    if template_count > 0:
        logger.info(f"Excluded {template_count} template/placeholder deals")

    # Apply deal name filter if specified
    if deal_filter:
        real_deals = [
            d for d in real_deals if deal_filter.lower() in d.deal_name.lower()
        ]
        logger.info(f"Filtered to {len(real_deals)} deals matching '{deal_filter}'")

    if not real_deals:
        logger.warning("No deals to audit!")
        return []

    # Batch-fetch all extracted values for all audited fields
    prop_ids = [d.property_id for d in real_deals]
    fields_param = ",".join(f"'{f}'" for f in ALL_AUDIT_FIELDS)
    prop_ids_param = ",".join(str(pid) for pid in prop_ids)

    ev_query = text(f"""
        SELECT ev.property_id, ev.field_name, ev.value_text, ev.value_numeric,
               ev.is_error, ev.error_category, ev.source_file, ev.sheet_name,
               ev.cell_address, ev.confidence_score
        FROM extracted_values ev
        INNER JOIN (
            SELECT ev2.property_id, MAX(er.completed_at) as max_completed
            FROM extracted_values ev2
            JOIN extraction_runs er ON ev2.extraction_run_id = er.id
            WHERE er.status = 'completed'
              AND ev2.property_id IN ({prop_ids_param})
            GROUP BY ev2.property_id
        ) latest ON ev.property_id = latest.property_id
        JOIN extraction_runs er2 ON ev.extraction_run_id = er2.id
            AND er2.completed_at = latest.max_completed
        WHERE ev.property_id IN ({prop_ids_param})
          AND ev.field_name IN ({fields_param})
    """)
    ev_rows = db.execute(ev_query).fetchall()

    # Build lookup: {property_id: {field_name: row_data}}
    ev_lookup: dict[int, dict[str, dict]] = defaultdict(dict)
    for row in ev_rows:
        pid, fname = row[0], row[1]
        # Keep first occurrence per (property_id, field_name) since we ordered by latest
        if fname not in ev_lookup[pid]:
            ev_lookup[pid][fname] = {
                "value_text": row[2],
                "value_numeric": float(row[3]) if row[3] is not None else None,
                "is_error": row[4],
                "error_category": row[5],
                "source_file": row[6],
                "sheet_name": row[7],
                "cell_address": row[8],
                "confidence_score": float(row[9]) if row[9] is not None else None,
            }

    logger.info(
        f"Fetched extracted values for {len(ev_lookup)} properties "
        f"({len(ev_rows)} total rows)"
    )

    # Classify each (deal, field) pair
    audit_results: list[FieldAudit] = []
    for deal in real_deals:
        fields = ev_lookup.get(deal.property_id, {})
        for field_name in ALL_AUDIT_FIELDS:
            ev = fields.get(field_name)
            if ev is None:
                # No extracted_value row exists at all
                audit_results.append(
                    FieldAudit(
                        deal_id=deal.deal_id,
                        deal_name=deal.deal_name,
                        property_id=deal.property_id,
                        field_name=field_name,
                        status="missing",
                        dashboard_shows="N/A",
                    )
                )
                continue

            vtext = ev["value_text"]
            vnum = ev["value_numeric"]
            is_err = ev["is_error"]

            if is_err:
                status = "error"
                dashboard_shows = "N/A"
            elif (
                vtext
                and vtext.strip().upper()
                in (
                    "N/A",
                    "NA",
                    "#N/A",
                    "TBD",
                    "TBA",
                    "-",
                    "NONE",
                )
                and vnum is None
            ):
                status = "n_a_placeholder"
                dashboard_shows = "N/A"
            elif vnum is not None:
                # Check enrichment guard conditions
                if field_name in POSITIVE_GUARD_FIELDS and vnum <= 0:
                    status = "guarded_zero"
                    dashboard_shows = (
                        "N/A (value exists but <=0, skipped by enrichment)"
                    )
                else:
                    status = "has_value"
                    dashboard_shows = str(vnum)
            elif vtext and vtext.strip():
                status = "has_value"
                dashboard_shows = vtext.strip()
            else:
                # Both null, not error
                status = "missing"
                dashboard_shows = "N/A"

            audit_results.append(
                FieldAudit(
                    deal_id=deal.deal_id,
                    deal_name=deal.deal_name,
                    property_id=deal.property_id,
                    field_name=field_name,
                    status=status,
                    value_text=vtext,
                    value_numeric=vnum,
                    is_error=is_err,
                    error_category=ev["error_category"],
                    source_file=ev["source_file"],
                    sheet_name=ev["sheet_name"],
                    cell_address=ev["cell_address"],
                    confidence_score=ev["confidence_score"],
                    dashboard_shows=dashboard_shows,
                )
            )

    # Summary stats
    status_counts = defaultdict(int)
    for r in audit_results:
        status_counts[r.status] += 1

    total = len(audit_results)
    logger.info(
        f"Phase 1 complete: {total} field checks across {len(real_deals)} deals"
    )
    for status, count in sorted(status_counts.items()):
        pct = count / total * 100 if total else 0
        logger.info(f"  {status}: {count} ({pct:.1f}%)")

    return audit_results


def write_completeness_report(results: list[FieldAudit]) -> Path:
    """Write the completeness report CSV."""
    out_path = REPORTS_DIR / "completeness_report.csv"
    fieldnames = [
        "deal_name",
        "deal_id",
        "property_id",
        "field_name",
        "status",
        "value_text",
        "value_numeric",
        "is_error",
        "error_category",
        "source_file",
        "sheet_name",
        "cell_address",
        "confidence_score",
        "dashboard_shows",
    ]
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for r in results:
            writer.writerow(
                {
                    "deal_name": r.deal_name,
                    "deal_id": r.deal_id,
                    "property_id": r.property_id,
                    "field_name": r.field_name,
                    "status": r.status,
                    "value_text": r.value_text or "",
                    "value_numeric": r.value_numeric
                    if r.value_numeric is not None
                    else "",
                    "is_error": r.is_error,
                    "error_category": r.error_category or "",
                    "source_file": Path(r.source_file).name if r.source_file else "",
                    "sheet_name": r.sheet_name or "",
                    "cell_address": r.cell_address or "",
                    "confidence_score": r.confidence_score
                    if r.confidence_score is not None
                    else "",
                    "dashboard_shows": r.dashboard_shows,
                }
            )
    logger.info(f"Wrote completeness report: {out_path} ({len(results)} rows)")
    return out_path


# ── Phase 2: Excel Source File Verification ───────────────────────────────────


def phase2_verify_sources(
    audit_results: list[FieldAudit],
) -> list[CellVerification]:
    """Verify N/A/error/missing values against actual Excel source files.

    Opens each source file once and reads all needed cells in batch.
    """
    logger.info("Phase 2: Starting Excel source file verification...")

    # Collect fields that need verification
    needs_verify = [
        r
        for r in audit_results
        if r.status in ("n_a_placeholder", "error", "missing", "guarded_zero")
    ]
    logger.info(f"  {len(needs_verify)} fields need source verification")

    if not needs_verify:
        logger.info("  No fields need verification!")
        return []

    # Load groups.json for fallback cell lookups
    groups_data = load_groups_json()

    # Group verification tasks by source file
    by_file: dict[str, list[FieldAudit]] = defaultdict(list)
    no_source: list[FieldAudit] = []

    for r in needs_verify:
        if r.source_file and Path(r.source_file).exists():
            by_file[r.source_file].append(r)
        elif r.source_file:
            # Try to find in UW model dir by filename
            fname = Path(r.source_file).name
            local_path = UW_MODEL_DIR / fname
            if local_path.exists():
                by_file[str(local_path)].append(r)
            else:
                no_source.append(r)
        else:
            no_source.append(r)

    if no_source:
        logger.warning(f"  {len(no_source)} fields have no accessible source file")

    logger.info(f"  Verifying cells across {len(by_file)} source files...")

    verifications: list[CellVerification] = []

    # For fields with no source file, try to find via groups.json
    for r in no_source:
        verifications.append(
            CellVerification(
                deal_name=r.deal_name,
                field_name=r.field_name,
                source_file="UNKNOWN",
                sheet_name=r.sheet_name or "",
                cell_address=r.cell_address or "",
                db_status=r.status,
                db_value_text=r.value_text,
                db_value_numeric=r.value_numeric,
                actual_cell_value=None,
                actual_cell_type="no_source_file",
                verdict="no_source_file",
            )
        )

    # Process each file
    file_count = 0
    for file_path_str, field_audits in by_file.items():
        file_path = Path(file_path_str)
        file_count += 1
        if file_count % 50 == 0:
            logger.info(f"  Processing file {file_count}/{len(by_file)}...")

        # Collect all cells to read from this file
        cells_to_read: list[tuple[str, str]] = []
        audit_for_cell: dict[tuple[str, str], FieldAudit] = {}

        for r in field_audits:
            sheet = r.sheet_name
            cell = r.cell_address
            if sheet and cell:
                cells_to_read.append((sheet, cell))
                audit_for_cell[(sheet, cell)] = r
            else:
                # Try to find expected cell from reference mapping
                _, ref_mapping = find_file_group(file_path_str, groups_data)
                cell_info = get_field_cell_from_mapping(ref_mapping, r.field_name)
                if cell_info:
                    cells_to_read.append(cell_info)
                    audit_for_cell[cell_info] = r
                else:
                    verifications.append(
                        CellVerification(
                            deal_name=r.deal_name,
                            field_name=r.field_name,
                            source_file=file_path.name,
                            sheet_name=sheet or "",
                            cell_address=cell or "",
                            db_status=r.status,
                            db_value_text=r.value_text,
                            db_value_numeric=r.value_numeric,
                            actual_cell_value=None,
                            actual_cell_type="no_cell_mapping",
                            verdict="no_cell_mapping",
                        )
                    )

        if not cells_to_read:
            continue

        # Read all cells from this file in one batch
        if file_path.suffix.lower() == ".xlsb":
            cell_values = read_xlsb_cells_batch(file_path, cells_to_read)
        else:
            # openpyxl for .xlsx
            cell_values = _read_xlsx_cells_batch(file_path, cells_to_read)

        # Classify each result
        for (sheet, cell), actual_value in cell_values.items():
            r = audit_for_cell.get((sheet, cell))
            if r is None:
                continue

            cell_type, display = classify_cell_value(actual_value)

            # Determine verdict
            if r.status in ("n_a_placeholder", "missing") and cell_type in (
                "empty",
                "n_a_text",
            ):
                verdict = "correct_na"
            elif r.status == "error" and cell_type == "formula_error":
                verdict = "correct_error"
            elif r.status == "error" and cell_type in ("empty", "n_a_text"):
                verdict = "correct_na"
            elif r.status == "guarded_zero" and cell_type == "numeric":
                # guarded_zero: DB has a value but enrichment >0 guard drops it
                try:
                    db_val = r.value_numeric or 0.0
                    actual_val = (
                        float(actual_value) if actual_value is not None else 0.0
                    )
                except (ValueError, TypeError):
                    actual_val = 0.0
                if actual_val > 0 and db_val <= 0:
                    # Excel has positive value but DB has 0/negative — real extraction error
                    verdict = "discrepancy"
                elif actual_val <= 0 and db_val <= 0:
                    # Both zero or negative — guard is working correctly
                    verdict = "correct_guard"
                else:
                    verdict = "correct_guard"
            elif cell_type == "numeric":
                verdict = "discrepancy"
            elif cell_type == "string" and display and not display.startswith("__"):
                # String value exists — may or may not be meaningful
                if r.field_name in (
                    "CURRENT_OWNER",
                    "PROPERTY_CITY",
                    "SUBMARKET",
                    "LAST_SALE_DATE",
                    "PROPERTY_ADDRESS",
                ):
                    verdict = "discrepancy"
                else:
                    # For numeric fields, a string value is likely a label, not data
                    verdict = "correct_na"
            elif cell_type == "empty":
                verdict = "correct_na"
            else:
                verdict = "correct_na"

            verifications.append(
                CellVerification(
                    deal_name=r.deal_name,
                    field_name=r.field_name,
                    source_file=file_path.name,
                    sheet_name=sheet,
                    cell_address=cell,
                    db_status=r.status,
                    db_value_text=r.value_text,
                    db_value_numeric=r.value_numeric,
                    actual_cell_value=display or actual_value,
                    actual_cell_type=cell_type,
                    verdict=verdict,
                )
            )

    # Summary
    verdict_counts = defaultdict(int)
    for v in verifications:
        verdict_counts[v.verdict] += 1

    logger.info(f"Phase 2 complete: {len(verifications)} cells verified")
    for verdict, count in sorted(verdict_counts.items()):
        logger.info(f"  {verdict}: {count}")

    return verifications


def _read_xlsx_cells_batch(
    file_path: Path, cells: list[tuple[str, str]]
) -> dict[tuple[str, str], Any]:
    """Read multiple cells from a .xlsx file using openpyxl."""
    import openpyxl

    results: dict[tuple[str, str], Any] = {}
    try:
        wb = openpyxl.load_workbook(str(file_path), data_only=True, read_only=True)
        for sheet_name, cell_address in cells:
            if sheet_name not in wb.sheetnames:
                results[(sheet_name, cell_address)] = f"__SHEET_MISSING:{sheet_name}__"
                continue
            ws = wb[sheet_name]
            try:
                cell = ws[cell_address]
                results[(sheet_name, cell_address)] = cell.value
            except Exception:
                results[(sheet_name, cell_address)] = None
        wb.close()
    except Exception as e:
        for sheet_name, cell_address in cells:
            if (sheet_name, cell_address) not in results:
                results[(sheet_name, cell_address)] = f"__FILE_ERROR:{e}__"
    return results


def write_discrepancy_report(verifications: list[CellVerification]) -> Path:
    """Write the discrepancy report CSV (all verifications, not just discrepancies)."""
    out_path = REPORTS_DIR / "discrepancy_report.csv"
    fieldnames = [
        "deal_name",
        "field_name",
        "source_file",
        "sheet_name",
        "cell_address",
        "db_status",
        "db_value_text",
        "db_value_numeric",
        "actual_cell_value",
        "actual_cell_type",
        "verdict",
    ]
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for v in verifications:
            writer.writerow(
                {
                    "deal_name": v.deal_name,
                    "field_name": v.field_name,
                    "source_file": v.source_file,
                    "sheet_name": v.sheet_name,
                    "cell_address": v.cell_address,
                    "db_status": v.db_status,
                    "db_value_text": v.db_value_text or "",
                    "db_value_numeric": v.db_value_numeric
                    if v.db_value_numeric is not None
                    else "",
                    "actual_cell_value": str(v.actual_cell_value)
                    if v.actual_cell_value is not None
                    else "",
                    "actual_cell_type": v.actual_cell_type,
                    "verdict": v.verdict,
                }
            )
    discrepancies = [v for v in verifications if v.verdict == "discrepancy"]
    logger.info(
        f"Wrote discrepancy report: {out_path} "
        f"({len(verifications)} total, {len(discrepancies)} discrepancies)"
    )
    return out_path


# ── Phase 3: Fix Discrepancies ────────────────────────────────────────────────


def phase3_fix_discrepancies(
    db: Session,
    verifications: list[CellVerification],
    audit_results: list[FieldAudit],
    dry_run: bool = True,
) -> list[dict]:
    """Fix extracted values where source has real data but DB has N/A/error.

    Creates a new ExtractionRun with trigger_type='audit_fix' for traceability.
    """
    discrepancies = [v for v in verifications if v.verdict == "discrepancy"]
    logger.info(
        f"Phase 3: {'DRY RUN - ' if dry_run else ''}"
        f"Processing {len(discrepancies)} discrepancies..."
    )

    if not discrepancies:
        logger.info("  No discrepancies to fix!")
        return []

    fix_log: list[dict] = []

    if not dry_run:
        # Create audit_fix extraction run
        run_id = str(uuid4())
        db.execute(
            text("""
                INSERT INTO extraction_runs
                    (id, started_at, status, trigger_type, files_discovered,
                     files_processed, files_failed, created_at, updated_at)
                VALUES
                    (:id, :now, 'completed', 'audit_fix', 0, 0, 0, :now, :now)
            """),
            {"id": run_id, "now": datetime.now(UTC)},
        )
        db.flush()
        logger.info(f"  Created audit_fix extraction run: {run_id}")

    # Build lookup for audit results to get property_id
    audit_lookup: dict[tuple[str, str], FieldAudit] = {}
    for r in audit_results:
        audit_lookup[(r.deal_name, r.field_name)] = r

    fixed_count = 0
    for v in discrepancies:
        audit = audit_lookup.get((v.deal_name, v.field_name))
        if not audit:
            continue

        actual = v.actual_cell_value
        cell_type = v.actual_cell_type

        # Determine new values
        new_numeric = None
        new_text = str(actual) if actual is not None else None
        if cell_type == "numeric":
            try:
                new_numeric = float(actual)
            except (ValueError, TypeError):
                pass

        fix_entry = {
            "deal_name": v.deal_name,
            "field_name": v.field_name,
            "property_id": audit.property_id,
            "old_status": v.db_status,
            "old_value_text": v.db_value_text,
            "old_value_numeric": v.db_value_numeric,
            "new_value_text": new_text,
            "new_value_numeric": new_numeric,
            "source_file": v.source_file,
            "sheet_name": v.sheet_name,
            "cell_address": v.cell_address,
            "fix_timestamp": datetime.now(UTC).isoformat(),
            "applied": not dry_run,
        }

        if not dry_run:
            # Find the property_name from the existing extracted_value
            prop_name_row = db.execute(
                text("""
                    SELECT DISTINCT property_name FROM extracted_values
                    WHERE property_id = :pid AND field_name = :fname
                    LIMIT 1
                """),
                {"pid": audit.property_id, "fname": v.field_name},
            ).fetchone()
            property_name = prop_name_row[0] if prop_name_row else v.deal_name

            # Upsert the corrected value
            db.execute(
                text("""
                    INSERT INTO extracted_values
                        (id, extraction_run_id, property_id, property_name,
                         field_name, value_text, value_numeric, is_error,
                         sheet_name, cell_address, source_file, created_at, updated_at)
                    VALUES
                        (:id, :run_id, :pid, :pname, :fname, :vtext, :vnum,
                         false, :sheet, :cell, :sfile, :now, :now)
                    ON CONFLICT (extraction_run_id, property_name, field_name)
                    DO UPDATE SET
                        value_text = EXCLUDED.value_text,
                        value_numeric = EXCLUDED.value_numeric,
                        is_error = false,
                        updated_at = EXCLUDED.updated_at
                """),
                {
                    "id": str(uuid4()),
                    "run_id": run_id,
                    "pid": audit.property_id,
                    "pname": property_name,
                    "fname": v.field_name,
                    "vtext": new_text,
                    "vnum": new_numeric,
                    "sheet": v.sheet_name,
                    "cell": v.cell_address,
                    "sfile": v.source_file,
                    "now": datetime.now(UTC),
                },
            )
            fixed_count += 1

        fix_log.append(fix_entry)

    if not dry_run and fixed_count > 0:
        # Update extraction run stats
        db.execute(
            text("""
                UPDATE extraction_runs
                SET files_processed = :count, completed_at = :now, updated_at = :now
                WHERE id = :id
            """),
            {"count": fixed_count, "now": datetime.now(UTC), "id": run_id},
        )
        db.commit()
        logger.info(f"  Fixed {fixed_count} extracted values in DB")
    elif not dry_run:
        db.rollback()

    action = "would fix" if dry_run else "fixed"
    logger.info(f"Phase 3 complete: {action} {len(fix_log)} discrepancies")
    return fix_log


def write_fix_log(fix_log: list[dict]) -> Path:
    """Write the fix log CSV."""
    out_path = REPORTS_DIR / "fix_log.csv"
    if not fix_log:
        out_path.write_text("No discrepancies to fix.\n")
        return out_path

    fieldnames = list(fix_log[0].keys())
    with open(out_path, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(fix_log)
    logger.info(f"Wrote fix log: {out_path} ({len(fix_log)} entries)")
    return out_path


# ── Phase 4: N/A Documentation & Summary ─────────────────────────────────────


def phase4_generate_reports(
    audit_results: list[FieldAudit],
    verifications: list[CellVerification],
    fix_log: list[dict],
) -> None:
    """Generate the N/A documentation CSV and summary report."""
    logger.info("Phase 4: Generating N/A documentation and summary...")

    # ── N/A Documentation ──
    na_docs: list[dict] = []

    # From Phase 2 verified N/As
    for v in verifications:
        if v.verdict in ("correct_na", "correct_error"):
            na_docs.append(
                {
                    "deal_name": v.deal_name,
                    "uw_model_file_name": v.source_file,
                    "sheet_name": v.sheet_name,
                    "cell_address": v.cell_address,
                    "field_name": v.field_name,
                    "field_description": FIELD_DESCRIPTIONS.get(
                        v.field_name, v.field_name
                    ),
                    "actual_cell_content": str(v.actual_cell_value)
                    if v.actual_cell_value
                    else "(empty)",
                    "cell_type": v.actual_cell_type,
                    "verified_date": datetime.now(UTC).strftime("%Y-%m-%d"),
                }
            )

    # From Phase 1 missing fields (no source to verify, but still N/A on dashboard)
    verified_keys = {(v.deal_name, v.field_name) for v in verifications}
    for r in audit_results:
        if (
            r.status in ("n_a_placeholder", "error", "missing")
            and (r.deal_name, r.field_name) not in verified_keys
        ):
            na_docs.append(
                {
                    "deal_name": r.deal_name,
                    "uw_model_file_name": Path(r.source_file).name
                    if r.source_file
                    else "UNKNOWN",
                    "sheet_name": r.sheet_name or "",
                    "cell_address": r.cell_address or "",
                    "field_name": r.field_name,
                    "field_description": FIELD_DESCRIPTIONS.get(
                        r.field_name, r.field_name
                    ),
                    "actual_cell_content": "(not verified - no source access)",
                    "cell_type": "unverified",
                    "verified_date": datetime.now(UTC).strftime("%Y-%m-%d"),
                }
            )

    # Write N/A documentation CSV
    na_path = REPORTS_DIR / "na_documentation.csv"
    if na_docs:
        with open(na_path, "w", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=list(na_docs[0].keys()))
            writer.writeheader()
            writer.writerows(na_docs)
    else:
        na_path.write_text("No N/A values found.\n")
    logger.info(f"Wrote N/A documentation: {na_path} ({len(na_docs)} entries)")

    # ── Summary Report ──
    total_deals = len({r.deal_id for r in audit_results})
    total_fields = len(audit_results)

    status_counts = defaultdict(int)
    for r in audit_results:
        status_counts[r.status] += 1

    verdict_counts = defaultdict(int)
    for v in verifications:
        verdict_counts[v.verdict] += 1

    # Top fields by N/A count
    na_by_field = defaultdict(int)
    for r in audit_results:
        if r.status in ("n_a_placeholder", "error", "missing"):
            na_by_field[r.field_name] += 1
    top_na_fields = sorted(na_by_field.items(), key=lambda x: -x[1])[:20]

    # Top deals by missing field count
    missing_by_deal = defaultdict(int)
    for r in audit_results:
        if r.status in ("n_a_placeholder", "error", "missing"):
            missing_by_deal[r.deal_name] += 1
    top_missing_deals = sorted(missing_by_deal.items(), key=lambda x: -x[1])[:20]

    # Deals with complete enrichment data (all 33 core fields have values)
    enrichment_by_deal: dict[str, dict[str, int]] = defaultdict(
        lambda: defaultdict(int)
    )
    for r in audit_results:
        if r.field_name in ENRICHMENT_FIELDS:
            enrichment_by_deal[r.deal_name][r.status] += 1
    complete_deals = sum(
        1
        for d, counts in enrichment_by_deal.items()
        if counts.get("has_value", 0) == len(ENRICHMENT_FIELDS)
    )

    summary = f"""# Dashboard Deal Data Audit — Summary Report

**Generated**: {datetime.now(UTC).strftime("%Y-%m-%d %H:%M:%S UTC")}

---

## Scope

| Metric | Value |
|--------|-------|
| Active deals audited | {total_deals} |
| Fields checked per deal | {len(ALL_AUDIT_FIELDS)} ({len(ENRICHMENT_FIELDS)} enrichment + {len(PROFORMA_FIELDS)} proforma) |
| Total field checks | {total_fields:,} |
| Deals with complete enrichment (all 33 fields) | {complete_deals} / {total_deals} |

## Field Status Breakdown

| Status | Count | % |
|--------|------:|--:|
"""
    for status in ["has_value", "n_a_placeholder", "error", "missing", "guarded_zero"]:
        count = status_counts.get(status, 0)
        pct = count / total_fields * 100 if total_fields else 0
        summary += f"| {status} | {count:,} | {pct:.1f}% |\n"

    if verifications:
        summary += f"""
## Source File Verification Results

| Verdict | Count |
|---------|------:|
"""
        for verdict in [
            "correct_na",
            "correct_error",
            "correct_guard",
            "discrepancy",
            "no_source_file",
            "no_cell_mapping",
        ]:
            count = verdict_counts.get(verdict, 0)
            if count:
                summary += f"| {verdict} | {count:,} |\n"

    if fix_log:
        applied = sum(1 for f in fix_log if f.get("applied"))
        summary += f"""
## Discrepancy Fixes

| Metric | Value |
|--------|------:|
| Total discrepancies found | {len(fix_log)} |
| Fixes applied to DB | {applied} |
| Dry-run only (not applied) | {len(fix_log) - applied} |
"""

    summary += f"""
## Top 20 Fields by N/A Count

| Field | N/A Count | Description |
|-------|----------:|-------------|
"""
    for fname, count in top_na_fields:
        desc = FIELD_DESCRIPTIONS.get(fname, "")
        summary += f"| {fname} | {count} | {desc} |\n"

    summary += f"""
## Top 20 Deals by Missing Fields

| Deal | Missing Fields |
|------|---------------:|
"""
    for dname, count in top_missing_deals:
        summary += f"| {dname} | {count} |\n"

    summary += """
---

## Dashboard Pages Covered

- **Kanban Board**: units, submarket, t12_cap_on_pp, basis_per_unit
- **Deal Detail Modal**: All 33 enrichment fields (units, avg SF, loss factors, NOI margin, basis, cap rates, debt/equity, returns, exit metrics, year built, lat/lng)
- **Deal Comparison**: Same 33 fields formatted side-by-side
- **Proforma Returns**: 30 year-specific fields (IRR/MOIC by year, NOI/unit, cap rates, DSCR, debt yield)

## Output Files

- `completeness_report.csv` — Per-deal x per-field status matrix
- `discrepancy_report.csv` — Source file verification results
- `na_documentation.csv` — All validated N/A values with source cell details
- `fix_log.csv` — Log of DB corrections applied (or would-be-applied in dry-run)
- `summary.md` — This file
"""

    summary_path = REPORTS_DIR / "summary.md"
    summary_path.write_text(summary)
    logger.info(f"Wrote summary report: {summary_path}")


# ── Main ──────────────────────────────────────────────────────────────────────


def main():
    parser = argparse.ArgumentParser(description="Dashboard Deal Data Audit")
    parser.add_argument(
        "--fix",
        action="store_true",
        help="Apply fixes for discrepancies (default: dry-run only)",
    )
    parser.add_argument(
        "--deal",
        type=str,
        default=None,
        help="Filter to a specific deal name (substring match)",
    )
    parser.add_argument(
        "--db-only",
        action="store_true",
        help="Skip Excel source file verification (Phase 2)",
    )
    args = parser.parse_args()

    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    logger.info("=" * 70)
    logger.info("Dashboard Deal Data Accuracy & Completeness Audit")
    logger.info(f"  Mode: {'LIVE FIX' if args.fix else 'DRY RUN'}")
    if args.deal:
        logger.info(f"  Filter: {args.deal}")
    if args.db_only:
        logger.info("  Skipping Excel verification (--db-only)")
    logger.info("=" * 70)

    db = SessionLocal()
    try:
        # Phase 1: DB audit
        audit_results = phase1_db_audit(db, deal_filter=args.deal)
        write_completeness_report(audit_results)

        # Phase 2: Source verification
        verifications: list[CellVerification] = []
        if not args.db_only:
            verifications = phase2_verify_sources(audit_results)
            write_discrepancy_report(verifications)
        else:
            logger.info("Phase 2: SKIPPED (--db-only)")

        # Phase 3: Fix discrepancies
        fix_log: list[dict] = []
        if verifications:
            fix_log = phase3_fix_discrepancies(
                db, verifications, audit_results, dry_run=not args.fix
            )
            write_fix_log(fix_log)
        else:
            logger.info("Phase 3: SKIPPED (no verifications)")

        # Phase 4: Reports
        phase4_generate_reports(audit_results, verifications, fix_log)

    finally:
        db.close()

    logger.info("=" * 70)
    logger.info("Audit complete! Reports written to:")
    logger.info(f"  {REPORTS_DIR}/")
    logger.info("=" * 70)


if __name__ == "__main__":
    main()
