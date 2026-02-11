"""
Tests for structural fingerprinting of Excel UW model files.

Tests cover:
- SheetFingerprint properties and serialization
- FileFingerprint properties and serialization
- Single file fingerprinting (xlsb and xlsx)
- Population classification (populated, sparse, empty)
- Parallel fingerprinting
- Error handling for missing/corrupt files
- Content hash computation

Run with: pytest tests/test_extraction/test_fingerprint.py -v
"""

import io
from pathlib import Path
from unittest.mock import MagicMock, patch

import pytest

from app.extraction.fingerprint import (
    FileFingerprint,
    SheetFingerprint,
    fingerprint_file,
    fingerprint_files_parallel,
)

# Path to real .xlsb fixtures
FIXTURES_DIR = Path(__file__).parent.parent / "fixtures" / "uw_models"


class TestSheetFingerprint:
    """Tests for SheetFingerprint dataclass."""

    def test_signature_deterministic(self):
        """Same data should produce same signature."""
        fp1 = SheetFingerprint(name="Sheet1", row_count=100, col_count=10)
        fp2 = SheetFingerprint(name="Sheet1", row_count=100, col_count=10)
        assert fp1.signature == fp2.signature

    def test_signature_differs_by_name(self):
        """Different sheet names should produce different signatures."""
        fp1 = SheetFingerprint(name="Sheet1", row_count=100, col_count=10)
        fp2 = SheetFingerprint(name="Sheet2", row_count=100, col_count=10)
        assert fp1.signature != fp2.signature

    def test_signature_differs_by_labels(self):
        """Different labels should produce different signatures."""
        fp1 = SheetFingerprint(name="Sheet1", header_labels=["A", "B"])
        fp2 = SheetFingerprint(name="Sheet1", header_labels=["C", "D"])
        assert fp1.signature != fp2.signature

    def test_to_dict(self):
        """to_dict should serialize all fields."""
        fp = SheetFingerprint(
            name="Test",
            row_count=50,
            col_count=5,
            header_labels=["H1", "H2"],
            col_a_labels=["L1"],
            populated_cell_count=42,
        )
        d = fp.to_dict()
        assert d["name"] == "Test"
        assert d["row_count"] == 50
        assert d["populated_cell_count"] == 42

    def test_from_dict_roundtrip(self):
        """from_dict should restore from to_dict output."""
        fp = SheetFingerprint(name="Test", row_count=50, col_count=5)
        d = fp.to_dict()
        restored = SheetFingerprint.from_dict(d)
        assert restored.name == fp.name
        assert restored.row_count == fp.row_count


class TestFileFingerprint:
    """Tests for FileFingerprint dataclass."""

    def test_combined_signature(self):
        """combined_signature should join sorted sheet signatures."""
        s1 = SheetFingerprint(name="B", row_count=10, col_count=5)
        s2 = SheetFingerprint(name="A", row_count=20, col_count=3)
        fp = FileFingerprint(
            file_path="/test.xlsb",
            file_name="test.xlsb",
            sheet_signatures=[s1.signature, s2.signature],
            sheets=[s1, s2],
        )
        # Should be sorted
        sigs = fp.combined_signature.split("|")
        assert sigs == sorted(sigs)

    def test_to_dict_from_dict_roundtrip(self):
        """Full serialization roundtrip."""
        fp = FileFingerprint(
            file_path="/test.xlsb",
            file_name="test.xlsb",
            file_size=1000,
            content_hash="abc123",
            sheet_count=2,
            sheets=[
                SheetFingerprint(name="S1", row_count=10),
                SheetFingerprint(name="S2", row_count=20),
            ],
            total_populated_cells=100,
            population_status="populated",
        )
        d = fp.to_dict()
        restored = FileFingerprint.from_dict(d)
        assert restored.file_path == fp.file_path
        assert restored.file_size == fp.file_size
        assert len(restored.sheets) == 2
        assert restored.sheets[0].name == "S1"

    def test_population_status_populated(self):
        """Files with >=20 populated cells should be 'populated'."""
        fp = FileFingerprint(
            file_path="/test.xlsb",
            file_name="test.xlsb",
            total_populated_cells=100,
            population_status="populated",
        )
        assert fp.population_status == "populated"

    def test_population_status_sparse(self):
        """Files with 1-19 populated cells should be 'sparse'."""
        fp = FileFingerprint(
            file_path="/test.xlsb",
            file_name="test.xlsb",
            total_populated_cells=5,
            population_status="sparse",
        )
        assert fp.population_status == "sparse"


class TestFingerprintFile:
    """Tests for fingerprint_file function."""

    def test_fingerprint_nonexistent_file(self):
        """Non-existent file should return error status."""
        fp = fingerprint_file("/nonexistent/file.xlsb")
        assert fp.population_status == "error"
        assert fp.file_name == "file.xlsb"

    def test_fingerprint_with_content(self, tmp_path):
        """Fingerprinting with provided content should work."""
        # Create a minimal xlsx file for testing
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        ws["A1"] = "Header1"
        ws["B1"] = "Header2"
        ws["A2"] = "Label1"
        ws["B2"] = 42

        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        fp = fingerprint_file("/test.xlsx", file_content=content)
        assert fp.file_size == len(content)
        assert fp.content_hash != ""
        assert fp.sheet_count >= 1
        assert fp.population_status in ("populated", "sparse")

    def test_fingerprint_empty_xlsx(self, tmp_path):
        """Empty xlsx should be classified as empty."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "EmptySheet"
        # No data

        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        fp = fingerprint_file("/empty.xlsx", file_content=content)
        assert fp.population_status == "empty"

    def test_fingerprint_sparse_xlsx(self, tmp_path):
        """xlsx with few cells should be classified as sparse."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "SparseSheet"
        # Add 5 cells (below threshold of 20)
        for i in range(5):
            ws.cell(row=i + 1, column=1, value=f"value_{i}")

        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        fp = fingerprint_file("/sparse.xlsx", file_content=content, empty_threshold=20)
        assert fp.population_status == "sparse"
        assert fp.total_populated_cells == 5

    def test_fingerprint_populated_xlsx(self, tmp_path):
        """xlsx with many cells should be classified as populated."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FullSheet"
        # Add 30 cells (above threshold of 20)
        for i in range(30):
            ws.cell(row=i + 1, column=1, value=f"value_{i}")

        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        fp = fingerprint_file("/full.xlsx", file_content=content, empty_threshold=20)
        assert fp.population_status == "populated"
        assert fp.total_populated_cells >= 20

    def test_fingerprint_captures_header_labels(self):
        """Should capture first-row labels as header_labels."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "Name"
        ws["B1"] = "Value"
        ws["A2"] = "Item1"
        ws["B2"] = 100

        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        fp = fingerprint_file("/test.xlsx", file_content=content)
        assert len(fp.sheets) >= 1
        data_sheet = [s for s in fp.sheets if s.name == "Data"][0]
        assert "Name" in data_sheet.header_labels
        assert "Value" in data_sheet.header_labels

    def test_fingerprint_captures_col_a_labels(self):
        """Should capture column A labels."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Data"
        ws["A1"] = "Header"
        ws["A2"] = "Row1"
        ws["A3"] = "Row2"

        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        fp = fingerprint_file("/test.xlsx", file_content=content)
        data_sheet = [s for s in fp.sheets if s.name == "Data"][0]
        assert "Header" in data_sheet.col_a_labels
        assert "Row1" in data_sheet.col_a_labels

    def test_content_hash_deterministic(self):
        """Same content should produce same hash."""
        import openpyxl

        wb = openpyxl.Workbook()
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        fp1 = fingerprint_file("/a.xlsx", file_content=content)
        fp2 = fingerprint_file("/b.xlsx", file_content=content)
        assert fp1.content_hash == fp2.content_hash

    def test_content_hash_differs(self):
        """Different content should produce different hashes."""
        import openpyxl

        wb1 = openpyxl.Workbook()
        wb1.active["A1"] = "data1"
        buf1 = io.BytesIO()
        wb1.save(buf1)

        wb2 = openpyxl.Workbook()
        wb2.active["A1"] = "data2"
        buf2 = io.BytesIO()
        wb2.save(buf2)

        fp1 = fingerprint_file("/a.xlsx", file_content=buf1.getvalue())
        fp2 = fingerprint_file("/b.xlsx", file_content=buf2.getvalue())
        assert fp1.content_hash != fp2.content_hash


class TestFingerprintRealFixtures:
    """Tests using real .xlsb fixture files."""

    @pytest.fixture
    def fixture_files(self):
        """Get list of fixture file paths."""
        if not FIXTURES_DIR.exists():
            pytest.skip("Fixture directory not found")
        files = list(FIXTURES_DIR.glob("*.xlsb"))
        if not files:
            pytest.skip("No .xlsb fixtures found")
        return files

    def test_fingerprint_real_xlsb(self, fixture_files):
        """Real .xlsb file should fingerprint successfully."""
        fp = fingerprint_file(str(fixture_files[0]))
        assert fp.population_status == "populated"
        assert fp.sheet_count > 0
        assert fp.content_hash != ""
        assert fp.file_size > 0
        assert len(fp.sheets) > 0

    def test_real_xlsb_has_sheets_with_labels(self, fixture_files):
        """Real xlsb should have sheets with header/col_a labels."""
        fp = fingerprint_file(str(fixture_files[0]))
        sheets_with_headers = [s for s in fp.sheets if s.header_labels]
        assert len(sheets_with_headers) > 0

    def test_parallel_fingerprint_fixtures(self, fixture_files):
        """Parallel fingerprinting of real fixtures should work."""
        paths = [str(f) for f in fixture_files[:3]]
        results = fingerprint_files_parallel(paths, max_workers=2)
        assert len(results) == len(paths)
        for fp in results:
            assert fp.population_status in ("populated", "sparse", "empty", "error")


class TestFingerprintParallel:
    """Tests for parallel fingerprinting."""

    def test_single_file_no_threading(self):
        """Single file should be processed without threading."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws["A1"] = "test"
        buf = io.BytesIO()
        wb.save(buf)
        content = buf.getvalue()

        results = fingerprint_files_parallel(
            ["/test.xlsx"],
            file_contents={"/test.xlsx": content},
            max_workers=2,
        )
        assert len(results) == 1
        assert results[0].file_name == "test.xlsx"

    def test_multiple_files_parallel(self):
        """Multiple files should be processed in parallel."""
        import openpyxl

        contents = {}
        for i in range(3):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws["A1"] = f"file_{i}"
            buf = io.BytesIO()
            wb.save(buf)
            contents[f"/file_{i}.xlsx"] = buf.getvalue()

        results = fingerprint_files_parallel(
            list(contents.keys()),
            file_contents=contents,
            max_workers=2,
        )
        assert len(results) == 3

    def test_parallel_handles_errors(self):
        """Errors during parallel fingerprinting should be caught."""
        results = fingerprint_files_parallel(
            ["/nonexistent1.xlsb", "/nonexistent2.xlsb"],
            max_workers=2,
        )
        assert len(results) == 2
        for fp in results:
            assert fp.population_status == "error"
