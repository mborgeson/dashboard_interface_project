"""
Unit tests for the Excel export service.

Tests ExcelExportService functionality including:
- Property exports with formatting
- Deal exports with pipeline stages
- Analytics reports with charts
- Style creation and application
"""
import pytest
from io import BytesIO
from unittest.mock import MagicMock, patch

# Test with or without openpyxl
try:
    import openpyxl
    from openpyxl import Workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class TestExcelExportService:
    """Tests for the ExcelExportService class."""

    @pytest.fixture
    def service(self):
        """Create ExcelExportService instance."""
        from app.services.export_service import ExcelExportService
        return ExcelExportService()

    @pytest.fixture
    def sample_properties(self):
        """Sample property data for testing."""
        return [
            {
                "id": 1,
                "name": "Phoenix Apartments",
                "property_type": "multifamily",
                "address": "123 Main St",
                "city": "Phoenix",
                "state": "AZ",
                "market": "Phoenix Metro",
                "total_units": 100,
                "year_built": 2015,
                "occupancy_rate": 95.5,
                "avg_rent_per_unit": 1500.00,
                "noi": 1200000.00,
                "cap_rate": 6.5,
            },
            {
                "id": 2,
                "name": "Dallas Office",
                "property_type": "office",
                "address": "456 Commerce Ave",
                "city": "Dallas",
                "state": "TX",
                "market": "Dallas-Fort Worth",
                "total_units": None,
                "year_built": 2020,
                "occupancy_rate": 88.0,
                "avg_rent_per_unit": None,
                "noi": 2500000.00,
                "cap_rate": 5.8,
            },
        ]

    @pytest.fixture
    def sample_deals(self):
        """Sample deal data for testing."""
        return [
            {
                "id": 1,
                "name": "Deal #001",
                "deal_type": "acquisition",
                "stage": "underwriting",
                "priority": "high",
                "asking_price": 15000000,
                "offer_price": 14000000,
                "projected_irr": 18.5,
                "projected_coc": 8.0,
                "projected_equity_multiple": 2.1,
                "hold_period_years": 5,
                "source": "CBRE",
                "created_at": "2024-01-15",
                "updated_at": "2024-02-20",
            },
            {
                "id": 2,
                "name": "Deal #002",
                "deal_type": "development",
                "stage": "due_diligence",
                "priority": "medium",
                "asking_price": 25000000,
                "offer_price": 23000000,
                "projected_irr": 22.0,
                "projected_coc": 6.5,
                "projected_equity_multiple": 2.5,
                "hold_period_years": 7,
                "source": "JLL",
                "created_at": "2024-02-01",
                "updated_at": "2024-03-10",
            },
            {
                "id": 3,
                "name": "Deal #003",
                "deal_type": "acquisition",
                "stage": "closed",
                "priority": "high",
                "asking_price": 10000000,
                "offer_price": 9500000,
                "projected_irr": 15.0,
                "projected_coc": 9.0,
                "projected_equity_multiple": 1.8,
                "hold_period_years": 3,
                "source": "Marcus & Millichap",
                "created_at": "2023-11-01",
                "updated_at": "2024-01-05",
            },
        ]

    # ==================== Initialization Tests ====================

    def test_service_initialization(self, service):
        """Test ExcelExportService initializes correctly."""
        assert service._styles_created is False
        assert service._header_style is None
        assert service._currency_style is None
        assert service._percent_style is None
        assert service._date_style is None

    def test_color_constants_defined(self, service):
        """Test branding color constants are defined."""
        assert "primary" in service.COLORS
        assert "secondary" in service.COLORS
        assert "header_bg" in service.COLORS
        assert "success" in service.COLORS
        assert "danger" in service.COLORS

    def test_stage_colors_defined(self, service):
        """Test pipeline stage colors are defined."""
        assert "lead" in service.STAGE_COLORS
        assert "underwriting" in service.STAGE_COLORS
        assert "due_diligence" in service.STAGE_COLORS
        assert "closed" in service.STAGE_COLORS
        assert "dead" in service.STAGE_COLORS

    # ==================== Property Export Tests ====================

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_properties_returns_bytesio(self, service, sample_properties):
        """Test export_properties returns BytesIO buffer."""
        result = service.export_properties(sample_properties)
        assert isinstance(result, BytesIO)
        assert result.getvalue()  # Should have content

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_properties_creates_valid_excel(self, service, sample_properties):
        """Test export creates valid Excel file."""
        result = service.export_properties(sample_properties)

        # Load and verify workbook
        wb = openpyxl.load_workbook(result)
        assert "Properties" in wb.sheetnames

        ws = wb["Properties"]
        # Check headers exist
        assert ws.cell(row=1, column=1).value == "ID"
        assert ws.cell(row=1, column=2).value == "Name"

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_properties_with_analytics(self, service, sample_properties):
        """Test export includes analytics summary sheet."""
        result = service.export_properties(sample_properties, include_analytics=True)

        wb = openpyxl.load_workbook(result)
        assert "Analytics Summary" in wb.sheetnames

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_properties_without_analytics(self, service, sample_properties):
        """Test export excludes analytics when disabled."""
        result = service.export_properties(sample_properties, include_analytics=False)

        wb = openpyxl.load_workbook(result)
        assert "Analytics Summary" not in wb.sheetnames

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_properties_empty_list(self, service):
        """Test export handles empty property list."""
        result = service.export_properties([])

        wb = openpyxl.load_workbook(result)
        ws = wb["Properties"]
        # Should have headers but no data rows
        assert ws.cell(row=1, column=1).value == "ID"
        assert ws.cell(row=2, column=1).value is None

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_properties_data_formatting(self, service, sample_properties):
        """Test property data is formatted correctly."""
        result = service.export_properties(sample_properties)

        wb = openpyxl.load_workbook(result)
        ws = wb["Properties"]

        # Check first property data
        assert ws.cell(row=2, column=1).value == 1  # ID
        assert ws.cell(row=2, column=2).value == "Phoenix Apartments"  # Name

    # ==================== Deal Export Tests ====================

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_deals_returns_bytesio(self, service, sample_deals):
        """Test export_deals returns BytesIO buffer."""
        result = service.export_deals(sample_deals)
        assert isinstance(result, BytesIO)
        assert result.getvalue()

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_deals_creates_valid_excel(self, service, sample_deals):
        """Test deal export creates valid Excel file."""
        result = service.export_deals(sample_deals)

        wb = openpyxl.load_workbook(result)
        assert "Deals" in wb.sheetnames

        ws = wb["Deals"]
        assert ws.cell(row=1, column=1).value == "ID"
        assert ws.cell(row=1, column=4).value == "Stage"

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_deals_with_pipeline(self, service, sample_deals):
        """Test export includes pipeline summary sheet."""
        result = service.export_deals(sample_deals, include_pipeline=True)

        wb = openpyxl.load_workbook(result)
        assert "Pipeline Summary" in wb.sheetnames

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_deals_without_pipeline(self, service, sample_deals):
        """Test export excludes pipeline when disabled."""
        result = service.export_deals(sample_deals, include_pipeline=False)

        wb = openpyxl.load_workbook(result)
        assert "Pipeline Summary" not in wb.sheetnames

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_deals_stage_coloring(self, service, sample_deals):
        """Test deal stages have appropriate coloring."""
        result = service.export_deals(sample_deals)

        wb = openpyxl.load_workbook(result)
        ws = wb["Deals"]

        # Stage column (column 4) should have fill colors
        stage_cell = ws.cell(row=2, column=4)
        assert stage_cell.fill is not None

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_pipeline_summary_counts(self, service, sample_deals):
        """Test pipeline summary correctly counts stages."""
        result = service.export_deals(sample_deals, include_pipeline=True)

        wb = openpyxl.load_workbook(result)
        ws = wb["Pipeline Summary"]

        # Should have total row
        # Find the "Total" row and verify count
        for row in range(1, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == "Total":
                assert ws.cell(row=row, column=2).value == 3  # 3 deals

    # ==================== Analytics Report Tests ====================

    @pytest.fixture
    def sample_analytics_data(self):
        """Sample analytics data for testing."""
        return {
            "dashboard_metrics": {
                "portfolio_summary": {
                    "total_properties": 25,
                    "total_units": 2500,
                    "total_sf": 1500000,
                    "total_value": 250000000,
                    "avg_occupancy": 94.5,
                    "avg_cap_rate": 6.2,
                },
                "kpis": {
                    "ytd_noi_growth": 5.2,
                    "ytd_rent_growth": 3.8,
                    "deals_in_pipeline": 12,
                    "deals_closed_ytd": 4,
                    "capital_deployed_ytd": 45000000,
                },
            },
            "portfolio_analytics": {
                "performance": {
                    "total_return": 12.5,
                    "income_return": 6.5,
                    "appreciation_return": 6.0,
                    "benchmark_return": 10.0,
                    "alpha": 2.5,
                },
            },
            "deal_pipeline": {
                "funnel": {
                    "lead": 20,
                    "initial_review": 15,
                    "underwriting": 8,
                    "due_diligence": 4,
                    "under_contract": 2,
                    "closed": 5,
                },
                "conversion_rates": {
                    "lead_to_review": 75.0,
                    "review_to_underwriting": 53.3,
                    "underwriting_to_close": 25.0,
                },
            },
        }

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_analytics_report(self, service, sample_analytics_data):
        """Test analytics report export."""
        result = service.export_analytics_report(
            sample_analytics_data["dashboard_metrics"],
            sample_analytics_data["portfolio_analytics"],
            sample_analytics_data["deal_pipeline"],
        )

        assert isinstance(result, BytesIO)
        wb = openpyxl.load_workbook(result)

        # Should have multiple sheets
        assert "Dashboard Summary" in wb.sheetnames
        assert "Portfolio Performance" in wb.sheetnames
        assert "Deal Pipeline" in wb.sheetnames

    # ==================== Style Tests ====================

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_create_styles_only_once(self, service):
        """Test styles are only created once."""
        wb = Workbook()

        service._create_styles(wb)
        assert service._styles_created is True

        # Second call should not error
        service._create_styles(wb)
        assert service._styles_created is True

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_auto_width_columns(self, service):
        """Test column width auto-adjustment."""
        wb = Workbook()
        ws = wb.active

        ws.cell(row=1, column=1, value="Short")
        ws.cell(row=1, column=2, value="This is a much longer column header")

        service._auto_width_columns(ws)

        # Column B should be wider than A
        assert ws.column_dimensions["B"].width > ws.column_dimensions["A"].width

    # ==================== Edge Case Tests ====================

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_with_none_values(self, service):
        """Test export handles None values gracefully."""
        properties = [{
            "id": 1,
            "name": "Test Property",
            "property_type": "multifamily",
            "address": None,
            "city": None,
            "state": None,
            "market": None,
            "total_units": None,
            "year_built": None,
            "occupancy_rate": None,
            "avg_rent_per_unit": None,
            "noi": None,
            "cap_rate": None,
        }]

        # Should not raise exception
        result = service.export_properties(properties)
        assert result.getvalue()

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_with_zero_values(self, service):
        """Test export handles zero values correctly."""
        properties = [{
            "id": 1,
            "name": "Empty Property",
            "property_type": "land",
            "occupancy_rate": 0,
            "noi": 0,
            "cap_rate": 0,
        }]

        result = service.export_properties(properties)
        assert result.getvalue()

    # ==================== Resource Cleanup Tests ====================

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_returns_seekable_buffer(self, service, sample_properties):
        """Test export returns buffer positioned at start for reading."""
        result = service.export_properties(sample_properties)

        # Buffer should be at start
        assert result.tell() == 0

        # Should be readable
        content = result.read()
        assert len(content) > 0

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_workbook_closes_after_save(self, service, sample_properties):
        """Test workbook is properly saved and closed."""
        result = service.export_properties(sample_properties)

        # Result should be a complete, readable workbook
        wb = openpyxl.load_workbook(result)
        ws = wb.active

        # Verify we can read data (workbook was properly saved)
        assert ws.max_row > 0
        wb.close()

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_handles_large_dataset(self, service):
        """Test export handles large dataset without memory issues."""
        # Create 500 properties
        large_properties = [
            {
                "id": i,
                "name": f"Property {i}",
                "property_type": "multifamily",
                "address": f"{i} Main St",
                "city": "City",
                "state": "ST",
                "market": "Market",
                "total_units": i * 10,
                "year_built": 2000 + (i % 24),
                "occupancy_rate": 90 + (i % 10),
                "avg_rent_per_unit": 1000 + i,
                "noi": i * 10000,
                "cap_rate": 5 + (i % 5),
            }
            for i in range(500)
        ]

        result = service.export_properties(large_properties)

        # Verify it completed
        assert result.getvalue()

        # Verify row count
        wb = openpyxl.load_workbook(result)
        ws = wb["Properties"]
        assert ws.max_row == 501  # 500 data rows + 1 header
        wb.close()

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_export_with_special_characters(self, service):
        """Test export handles special characters in data."""
        properties = [{
            "id": 1,
            "name": "Test & Property <with> \"special\" chars",
            "property_type": "office",
            "address": "123 Main St, Suite #5",
            "city": "New York",
            "state": "NY",
            "market": "NYC Metro / Tri-State",
            "total_units": None,
            "year_built": 2020,
            "occupancy_rate": 95.5,
            "avg_rent_per_unit": None,
            "noi": 1000000.00,
            "cap_rate": 5.5,
        }]

        result = service.export_properties(properties)
        assert result.getvalue()

        # Verify special characters are preserved
        wb = openpyxl.load_workbook(result)
        ws = wb["Properties"]
        assert "Test & Property" in str(ws.cell(row=2, column=2).value)
        wb.close()

    @pytest.mark.skipif(not OPENPYXL_AVAILABLE, reason="openpyxl not installed")
    def test_deals_export_buffer_cleanup(self, service, sample_deals):
        """Test deals export properly cleans up resources."""
        result = service.export_deals(sample_deals)

        # Buffer should be seekable and readable
        result.seek(0)
        content = result.read()
        assert len(content) > 0

        # Should be reusable
        result.seek(0)
        wb = openpyxl.load_workbook(result)
        assert "Deals" in wb.sheetnames
        wb.close()

    def test_export_without_openpyxl_raises(self, service):
        """Test export raises ImportError when openpyxl unavailable."""
        with patch('app.services.export_service.OPENPYXL_AVAILABLE', False):
            from app.services.export_service import ExcelExportService
            svc = ExcelExportService()

            with pytest.raises(ImportError, match="openpyxl"):
                svc.export_properties([])


class TestExcelServiceSingleton:
    """Tests for the get_excel_service singleton."""

    def test_get_excel_service_returns_instance(self):
        """Test get_excel_service returns ExcelExportService instance."""
        from app.services.export_service import get_excel_service, ExcelExportService

        service = get_excel_service()
        assert isinstance(service, ExcelExportService)

    def test_get_excel_service_returns_same_instance(self):
        """Test get_excel_service returns singleton."""
        from app.services.export_service import get_excel_service

        service1 = get_excel_service()
        service2 = get_excel_service()
        assert service1 is service2
